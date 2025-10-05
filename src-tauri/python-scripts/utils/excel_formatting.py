"""
Utilities for Excel formatting and styling
"""
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def get_standard_styles() -> dict:
    """
    Retourne les styles Excel standards utilisés dans les rapports.
    
    Returns:
        Dictionnaire contenant tous les styles standards
    """
    return {
        'title_font': Font(bold=True, size=12),
        'header_font': Font(bold=True),
        'gray_fill': PatternFill("solid", fgColor="DDDDDD"),
        'center': Alignment(horizontal="center", vertical="center"),
        'center_wrap': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'black_thin': Side(style="thin", color="000000"),
        'border': None  # Sera créé dynamiquement
    }


def get_border(thin_side: Side = None) -> Border:
    """
    Crée une bordure complète avec des côtés fins noirs.
    
    Args:
        thin_side: Style de côté à utiliser (optionnel)
        
    Returns:
        Objet Border configuré
    """
    if thin_side is None:
        thin_side = Side(style="thin", color="000000")
    
    return Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )


def apply_cell_formatting(cell, value, cell_type: str = "text", is_special_row: bool = False,
                         styles: dict = None):
    """
    Applique le formatage standard à une cellule.

    Args:
        cell: Cellule openpyxl à formater
        value: Valeur à insérer dans la cellule
        cell_type: Type de formatage ("text", "number", "retention_time", "time", "header")
        is_special_row: Si True, applique le remplissage gris (pour lignes Total/Moyennes)
        styles: Dictionnaire des styles (optionnel)
    """
    if styles is None:
        styles = get_standard_styles()
        styles['border'] = get_border(styles['black_thin'])

    # Définir la valeur
    cell.value = value

    # Appliquer la bordure
    cell.border = styles['border']

    # Formatage spécifique selon le type
    if cell_type == "header":
        cell.font = styles['header_font']
        cell.fill = styles['gray_fill']
        cell.alignment = styles['center_wrap']
    elif cell_type == "retention_time":
        cell.number_format = "0.000"
        if is_special_row:
            cell.fill = styles['gray_fill']
    elif cell_type == "number":
        cell.number_format = "0.00"
        if is_special_row:
            cell.fill = styles['gray_fill']
    elif cell_type == "time":
        cell.alignment = styles['center']
        if is_special_row:
            cell.fill = styles['gray_fill']
    else:  # text
        if is_special_row:
            cell.fill = styles['gray_fill']


def set_column_widths(ws: Worksheet, column_configs: list[dict]):
    """
    Définit les largeurs des colonnes selon la configuration.
    
    Args:
        ws: Feuille de calcul openpyxl
        column_configs: Liste de configurations [{"column": 1, "width": 20}, ...]
    """
    for config in column_configs:
        col_letter = get_column_letter(config["column"])
        ws.column_dimensions[col_letter].width = config["width"]


def format_table_headers(ws: Worksheet, headers: list[str], start_row: int, start_col: int = 1,
                        styles: dict = None):
    """
    Formate les en-têtes d'un tableau avec le style standard.
    
    Args:
        ws: Feuille de calcul openpyxl
        headers: Liste des noms d'en-têtes
        start_row: Ligne de début des en-têtes
        start_col: Colonne de début (par défaut 1)
        styles: Dictionnaire des styles (optionnel)
    """
    if styles is None:
        styles = get_standard_styles()
        styles['border'] = get_border(styles['black_thin'])
    
    for j, header in enumerate(headers, start=start_col):
        # Formatage spécial pour les colonnes Rel. Area
        if header.startswith('Rel. Area (%) : '):
            element_name = header.replace('Rel. Area (%) : ', '')
            formatted_header = f"Rel. Area (%)\n{element_name}"
        # Ajouter l'unité (hh:mm) pour Injection Time
        elif header == 'Injection Time':
            formatted_header = f"Injection Time\n(hh:mm)"
        # Ajouter l'unité (min) pour RetentionTime
        elif header == 'RetentionTime' or header == 'Retention Time':
            formatted_header = "Retention Time\n(min)"
        # Ajouter l'unité (%) pour Relative Area
        elif header == 'Relative Area':
            formatted_header = "Relative Area\n(%)"
        else:
            formatted_header = header

        apply_cell_formatting(
            ws.cell(row=start_row, column=j),
            formatted_header,
            cell_type="header",
            styles=styles
        )


def format_data_table(ws: Worksheet, data_df, start_row: int, start_col: int = 1,
                     special_row_identifier: str = None, styles: dict = None):
    """
    Formate un tableau de données avec les styles standards.
    
    Args:
        ws: Feuille de calcul openpyxl
        data_df: DataFrame contenant les données
        start_row: Ligne de début des données
        start_col: Colonne de début (par défaut 1)
        special_row_identifier: Valeur qui identifie les lignes spéciales (ex: "Moyennes", "Total:")
        styles: Dictionnaire des styles (optionnel)
    """
    if styles is None:
        styles = get_standard_styles()
        styles['border'] = get_border(styles['black_thin'])
    
    headers = list(data_df.columns)
    
    for i, (_, row) in enumerate(data_df.iterrows()):
        r = start_row + i
        is_special = (special_row_identifier and 
                     str(row.iloc[0]).lower() == special_row_identifier.lower())
        
        for j, header in enumerate(headers, start=start_col):
            val = row[header]

            # Déterminer le type de cellule
            if j == 1:  # Première colonne (généralement nom)
                cell_type = "text"
            elif 'RetentionTime' in header or 'Retention Time' in header:
                cell_type = "retention_time"  # Format 3 décimales
            elif 'Injection Time' in header or (j == 2 and 'Time' in header):
                cell_type = "time"
            elif header.startswith('Rel. Area') or 'Relative Area' in header or '(%)' in header:
                cell_type = "number"  # Format 2 décimales pour pourcentages
            else:
                cell_type = "text"

            apply_cell_formatting(
                ws.cell(row=r, column=j),
                val,
                cell_type=cell_type,
                is_special_row=is_special,
                styles=styles
            )


def apply_standard_column_widths(ws: Worksheet, table_type: str = "main"):
    """
    Applique les largeurs de colonnes standards selon le type de tableau.
    
    Args:
        ws: Feuille de calcul openpyxl
        table_type: Type de tableau ("main", "summary", "carbon_family", "hvc")
    """
    if table_type == "main":
        # Tableau principal des données d'injection
        configs = [
            {"column": 1, "width": 20},  # Injection Name
            {"column": 2, "width": 16},  # Injection Time
        ]
        # Colonnes Rel. Area
        for i in range(3, 15):  # Jusqu'à 12 colonnes supplémentaires
            configs.append({"column": i, "width": 15})
    
    elif table_type == "summary":
        # Tableau de résumé (Peakname, RetentionTime, Relative Area)
        configs = [
            {"column": 1, "width": 25},  # Peakname
            {"column": 2, "width": 15},  # RetentionTime
            {"column": 3, "width": 15},  # Relative Area
        ]
    
    elif table_type == "carbon_family":
        # Tableau regroupement carbone/famille
        configs = [
            {"column": 1, "width": 12},  # Carbon
        ]
        # Colonnes familles
        for i in range(2, 8):
            configs.append({"column": i, "width": 14})
    
    elif table_type == "hvc":
        # Tableau HVC
        configs = [
            {"column": 1, "width": 22},  # Molécule
            {"column": 2, "width": 14},  # Moyenne (%)
        ]
    
    else:
        return  # Type inconnu
    
    set_column_widths(ws, configs)


def create_title_cell(ws: Worksheet, row: int, col: int, title: str, styles: dict = None):
    """
    Crée une cellule de titre avec le formatage standard.
    
    Args:
        ws: Feuille de calcul openpyxl
        row: Ligne de la cellule
        col: Colonne de la cellule
        title: Texte du titre
        styles: Dictionnaire des styles (optionnel)
    """
    if styles is None:
        styles = get_standard_styles()
    
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = styles['title_font']


def freeze_panes_standard(ws: Worksheet):
    """
    Fige les volets selon la configuration standard.
    
    Args:
        ws: Feuille de calcul openpyxl
    """
    ws.freeze_panes = "A3"