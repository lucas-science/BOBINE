import os
import sys
import pandas as pd
import traceback
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout

from utils.pignat.pignat_constants import (
    TIME,
    TT301,
    TT302,
    TT303,
    TT206,
    FT240,
    PI177,
    PT230,
    DATA_REQUIRED,
    GRAPHS,
    TEMPERATURE_DEPENDING_TIME,
    DEBIMETRIC_RESPONSE_DEPENDING_TIME,
    PRESSURE_PYROLYSEUR_DEPENDING_TIME,
    PRESSURE_POMPE_DEPENDING_TIME,
    DELTA_PRESSURE_DEPENDING_TIME,
    TEMPERATURE_DISPLAY_TITLE,
    DEBIMETRIC_DISPLAY_TITLE,
    PRESSURE_PYROLYSEUR_DISPLAY_TITLE,
    PRESSURE_POMPE_DISPLAY_TITLE,
    DELTA_PRESSURE_DISPLAY_TITLE,
    DISPLAY_NAME_MAPPING
)


class PignatData:
    def __init__(self, dir_root: str):
        self.first_file = ""
        if os.path.exists(dir_root):
            files = [f for f in os.listdir(dir_root)
                     if os.path.isfile(os.path.join(dir_root, f))
                     and not f.startswith('.')
                     and not f.startswith('~')
                     and not f.startswith('.~lock')
                     and f.lower().endswith('.csv')]

            if not files:
                raise FileNotFoundError(
                    f"Aucun fichier CSV valide trouvé dans {dir_root}")

            files.sort()
            self.first_file = os.path.join(dir_root, files[0])
        else:
            raise FileNotFoundError(f"Le répertoire {dir_root} n'existe pas")

        encodings_to_try = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252']
        separator = ','

        for encoding in encodings_to_try:
            try:
                with open(self.first_file, 'r', encoding=encoding) as f:
                    first_line = f.readline()
                    if first_line.count(';') > first_line.count(','):
                        separator = ';'
                    else:
                        separator = ','
                break
            except UnicodeDecodeError:
                if encoding == encodings_to_try[-1]:
                    raise UnicodeDecodeError(f"Failed to read file with any encoding: {encodings_to_try}")
                continue

        for encoding in encodings_to_try:
            try:
                self.data_frame = pd.read_csv(self.first_file, sep=separator, encoding=encoding)
                break
            except (UnicodeDecodeError, pd.errors.EmptyDataError, pd.errors.ParserError):
                if encoding == encodings_to_try[-1]:
                    raise ValueError(f"Failed to read CSV file with any encoding: {encodings_to_try}")
                continue

        self.columns = self.data_frame.columns.tolist()
        self.missing_columns = set(DATA_REQUIRED) - set(self.columns)

    def _select_columns(self, columns: list[str]) -> pd.DataFrame:
        missing_columns = [col for col in columns if col not in self.data_frame.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")

        return self.data_frame[columns]

    def _filter_by_time_range(self, df: pd.DataFrame, start_time=None, end_time=None) -> pd.DataFrame:
        if start_time is None and end_time is None:
            return df

        if TIME not in df.columns:
            return df

        sample_timestamp = df[TIME].iloc[0] if len(df) > 0 else None

        if sample_timestamp and ':' in str(sample_timestamp) and len(str(sample_timestamp).split()) == 1:
            mask = pd.Series([True] * len(df))

            if start_time is not None:
                try:
                    start_time_only = str(start_time).split(' ')[1] if ' ' in str(start_time) else str(start_time)
                    mask &= (df[TIME] >= start_time_only)
                except:
                    pass

            if end_time is not None:
                try:
                    end_time_only = str(end_time).split(' ')[1] if ' ' in str(end_time) else str(end_time)
                    mask &= (df[TIME] <= end_time_only)
                except:
                    pass
        else:
            mask = pd.Series([True] * len(df))
            if start_time is not None:
                mask &= (df[TIME] >= start_time)
            if end_time is not None:
                mask &= (df[TIME] <= end_time)

        filtered_df = df[mask]
        return filtered_df


    def is_all_required_data(self) -> bool:
        return all(col in self.data_frame.columns for col in DATA_REQUIRED)

    def get_available_graphs(self) -> list[dict]:
        graphs = []
        for graph in GRAPHS:
            graph_dict = {
                'name': graph['name'],  # Internal ID (ASCII, no accents)
                'displayName': DISPLAY_NAME_MAPPING.get(graph['name'], graph['name']),  # Beautiful name for UI
                'available': all(col in self.columns for col in graph['columns']),
                'columns': graph['columns']
            }
            graphs.append(graph_dict)
        return graphs

    def get_time_range(self) -> dict:
        if TIME not in self.columns:
            raise ValueError(f"Column {TIME} not found in data")
        
        time_column = self.data_frame[TIME]
        all_times = sorted(time_column.dropna().unique().tolist())

        if not all_times:
            return {
                "min_time": None,
                "max_time": None,
                "unique_times": []
            }

        min_time = all_times[0]
        max_time = all_times[-1]

        try:
            min_dt = pd.to_datetime(min_time)
            max_dt = pd.to_datetime(max_time)
            duration_minutes = (max_dt - min_dt).total_seconds() / 60

            target_points = 72
            delta_minutes = max(1, int(duration_minutes / target_points))
            delta = pd.Timedelta(minutes=delta_minutes)

            step_times = []
            current_time = min_dt
            while current_time <= max_dt:
                step_times.append(current_time.strftime('%Y-%m-%d %H:%M:%S'))
                current_time += delta

            return {
                "min_time": min_time,
                "max_time": max_time,
                "unique_times": step_times
            }

        except Exception:
            step = max(1, len(all_times) // 100)
            sampled_times = all_times[::step]
            return {
                "min_time": min_time,
                "max_time": max_time,
                "unique_times": sampled_times
            }


    def report_missing_per_column(self) -> pd.Series:
        return self.data_frame.isna().sum()

    def report_missing_per_row(self) -> pd.DataFrame:
        df = self.data_frame
        count_na = df.isna().sum(axis=1)
        cols_na = df.isna().apply(lambda row: list(row[row].index), axis=1)
        return pd.DataFrame({
            'n_missing': count_na,
            'cols_missing': cols_na
        }, index=df.index)


    def _get_temperature_over_time(self, start_time=None, end_time=None) -> pd.DataFrame:
        cols = [TIME, TT301, TT302, TT303, TT206]
        df = self._select_columns(cols).copy()
        return self._filter_by_time_range(df, start_time, end_time)

    def _get_debimetrique_response_over_time(self, start_time=None, end_time=None) -> pd.DataFrame:
        cols = [TIME, FT240]
        df = self._select_columns(cols).copy()
        return self._filter_by_time_range(df, start_time, end_time)

    def _get_pression_pyrolyseur_over_time(self, start_time=None, end_time=None) -> pd.DataFrame:
        cols = [TIME, PI177]
        df = self._select_columns(cols).copy()
        return self._filter_by_time_range(df, start_time, end_time)

    def _get_pression_sortie_pompe_over_time(self, start_time=None, end_time=None) -> pd.DataFrame:
        cols = [TIME, PT230]
        df = self._select_columns(cols).copy()
        return self._filter_by_time_range(df, start_time, end_time)

    def _get_delta_pression_over_time(self, start_time=None, end_time=None) -> pd.DataFrame:
        cols = [TIME, PI177, PT230]
        df_sel = self._select_columns(cols).copy()
        df_sel = self._filter_by_time_range(df_sel, start_time, end_time)
        delta_name = f"Delta_Pression_{PI177}_minus_{PT230}"
        df_sel[delta_name] = df_sel[PI177] - df_sel[PT230]
        return df_sel.drop(columns=[PI177, PT230])


    def get_json_metrics(self, metric: str, start_time=None, end_time=None):
        try:
            if metric == TEMPERATURE_DEPENDING_TIME:
                required_cols = [TIME, TT301, TT302, TT303, TT206]
                missing_cols = [col for col in required_cols if col not in self.columns]
                if missing_cols:
                    raise ValueError(f"Missing columns for temperature metric: {missing_cols}")

                return {
                    "name": TEMPERATURE_DISPLAY_TITLE,
                    "data": self._get_temperature_over_time(start_time, end_time),
                    "x_axis": TIME,
                    "y_axis": [TT301, TT302, TT303, TT206]
                }
            elif metric == DEBIMETRIC_RESPONSE_DEPENDING_TIME:
                required_cols = [TIME, FT240]
                missing_cols = [col for col in required_cols if col not in self.columns]
                if missing_cols:
                    raise ValueError(f"Missing columns for debimetric metric: {missing_cols}")

                return {
                    "name": DEBIMETRIC_DISPLAY_TITLE,
                    "data": self._get_debimetrique_response_over_time(start_time, end_time),
                    "x_axis": TIME,
                    "y_axis": [FT240]
                }
            elif metric == PRESSURE_PYROLYSEUR_DEPENDING_TIME:
                required_cols = [TIME, PI177]
                missing_cols = [col for col in required_cols if col not in self.columns]
                if missing_cols:
                    raise ValueError(f"Missing columns for pyrolyseur pressure metric: {missing_cols}")

                return {
                    "name": PRESSURE_PYROLYSEUR_DISPLAY_TITLE,
                    "data": self._get_pression_pyrolyseur_over_time(start_time, end_time),
                    "x_axis": TIME,
                    "y_axis": [PI177]
                }
            elif metric == PRESSURE_POMPE_DEPENDING_TIME:
                required_cols = [TIME, PT230]
                missing_cols = [col for col in required_cols if col not in self.columns]
                if missing_cols:
                    raise ValueError(f"Missing columns for pump pressure metric: {missing_cols}")

                return {
                    "name": PRESSURE_POMPE_DISPLAY_TITLE,
                    "data": self._get_pression_sortie_pompe_over_time(start_time, end_time),
                    "x_axis": TIME,
                    "y_axis": [PT230]
                }
            elif metric == DELTA_PRESSURE_DEPENDING_TIME:
                required_cols = [TIME, PI177, PT230]
                missing_cols = [col for col in required_cols if col not in self.columns]
                if missing_cols:
                    raise ValueError(f"Missing columns for delta pressure metric: {missing_cols}")

                return {
                    "name": DELTA_PRESSURE_DISPLAY_TITLE,
                    "data": self._get_delta_pression_over_time(start_time, end_time),
                    "x_axis": TIME,
                    "y_axis": [f"Delta_Pression_{PI177}_minus_{PT230}"]
                }
            else:
                raise ValueError(f"Metric '{metric}' is not recognized.")
        except Exception:
            raise


    def generate_workbook_with_charts(self,
        wb: Workbook,
        metrics_wanted: list,
        sheet_name: str = "Pignat"
    ) -> Workbook:
        
        ws = wb.create_sheet(title=sheet_name)
        
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_col = 1

        for i, metric_config in enumerate(metrics_wanted):
            try:
                if metric_config is None:
                    continue

                if isinstance(metric_config, dict):
                    metric_name = metric_config.get("name")
                    if metric_name is None:
                        continue
                    time_range = metric_config.get("timeRange", {})
                    start_time = time_range.get("startTime") if time_range else None
                    end_time = time_range.get("endTime") if time_range else None
                else:
                    metric_name = metric_config
                    start_time = None
                    end_time = None

                if not metric_name:
                    continue

                metric_data = self.get_json_metrics(metric_name, start_time, end_time)
                df = metric_data['data']

                # Resample to 1 point per minute for precise granularity
                if not df.empty and TIME in df.columns:
                    try:
                        df_copy = df.copy()
                        sample_time = str(df_copy[TIME].iloc[0])

                        if ':' in sample_time and len(sample_time.split()) == 1:
                            df_copy['_datetime'] = pd.to_datetime('2000-01-01 ' + df_copy[TIME].astype(str))
                        else:
                            df_copy['_datetime'] = pd.to_datetime(df_copy[TIME])

                        df_copy.set_index('_datetime', inplace=True)
                        numeric_cols = [col for col in df_copy.columns if col != TIME and pd.api.types.is_numeric_dtype(df_copy[col])]
                        df_resampled = df_copy[numeric_cols].resample('1T').mean().dropna(how='all')
                        df_resampled.reset_index(inplace=True)

                        if ':' in sample_time and len(sample_time.split()) == 1:
                            df_resampled[TIME] = df_resampled['_datetime'].dt.strftime('%H:%M:%S')
                        else:
                            df_resampled[TIME] = df_resampled['_datetime'].dt.strftime('%Y-%m-%d %H:%M:%S')

                        df_resampled.drop('_datetime', axis=1, inplace=True)
                        cols_order = [TIME] + [col for col in df_resampled.columns if col != TIME]
                        df_display = df_resampled[cols_order]
                    except Exception:
                        df_display = df.copy()
                else:
                    df_display = df.copy()

                df_table = df_display.copy()
                
                title = metric_data['name'].replace('=', '-')
                title_cell = ws.cell(row=1, column=current_col, value=title)
                title_cell.font = header_font
                
                for j, col_name in enumerate(df_table.columns):
                    header_cell = ws.cell(row=2, column=current_col + j, value=col_name)
                    header_cell.font = header_font
                    header_cell.border = thin_border
                    
                    if col_name == TIME:
                        col_letter = get_column_letter(current_col + j)
                        ws.column_dimensions[col_letter].width = 12
                
                for row_idx, row_data in enumerate(df_table.itertuples(index=False)):
                    for col_idx, value in enumerate(row_data):
                        data_cell = ws.cell(row=3 + row_idx, column=current_col + col_idx, value=value)
                        data_cell.border = thin_border
                
                chart = LineChart()
                chart.title = title
                chart.style = 2

                is_mono_series = len(metric_data['y_axis']) == 1

                chart.y_axis.title = (
                    ', '.join(metric_data['y_axis'])
                    if len(metric_data['y_axis']) > 1
                    else metric_data['y_axis'][0]
                )
                chart.x_axis.title = metric_data['x_axis']
                
                chart.x_axis.tickLblPos = "low"
                chart.y_axis.tickLblPos = "low"

                num_data_points = len(df_table)

                # Pour ~1440 points (24h à 1pt/min), viser 8-10 lignes de grille max
                if num_data_points > 720:  # > 12h
                    tick_interval = max(1, num_data_points // 24)  # ~24 labels
                    grid_interval = max(1, num_data_points // 8)   # ~8 lignes de grille
                elif num_data_points > 180:  # 3h-12h
                    tick_interval = max(1, num_data_points // 15)
                    grid_interval = max(1, num_data_points // 6)
                elif num_data_points > 60:   # 1h-3h
                    tick_interval = max(1, num_data_points // 10)
                    grid_interval = max(1, num_data_points // 4)
                else:
                    tick_interval = max(1, num_data_points // 5)
                    grid_interval = max(1, num_data_points // 3)

                chart.x_axis.tickLblSkip = max(0, tick_interval - 1) if tick_interval > 2 else 0
                chart.x_axis.tickMarkSkip = max(0, grid_interval - 1) if grid_interval > 1 else 0

                chart.y_axis.tickLblSkip = 0
                chart.y_axis.majorGridlines = ChartLines()
                chart.x_axis.majorGridlines = ChartLines()
                chart.x_axis.textRotation = -30
                chart.x_axis.delete = False
                chart.y_axis.delete = False
                
                if not is_mono_series:
                    chart.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode="edge",
                            yMode="edge",
                            x=0.10,
                            y=0.10,
                            w=0.80,
                            h=0.85
                        )
                    )
                else:
                    chart.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode="edge",
                            yMode="edge",
                            x=0.05,
                            y=0.10,
                            w=0.95,
                            h=0.85
                        )
                    )

                chart.width = 23
                chart.height = 13
                
                max_row_table = 2 + len(df_table)
                
                data_ref = Reference(ws,
                                    min_col=current_col + 1,
                                    min_row=2,
                                    max_col=current_col + len(metric_data['y_axis']),
                                    max_row=max_row_table)
                chart.add_data(data_ref, titles_from_data=True)

                if is_mono_series:
                    chart.legend = None

                    uniform_color = "1f77b4"

                    for series in chart.series:
                        try:
                            series.graphicalProperties.line.solidFill = uniform_color
                            series.graphicalProperties.line.width = 25000

                            if hasattr(series, 'marker'):
                                series.marker.symbol = "none"

                            series.smooth = True
                        except Exception:
                            pass
                else:
                    chart.legend.position = 'r'
                    chart.legend.overlay = False
                
                cats = Reference(ws,
                                min_col=current_col,
                                min_row=3,
                                max_row=max_row_table)
                chart.set_categories(cats)
                
                chart.x_axis.crosses = "min"
                chart.y_axis.crosses = "min"

                try:
                    if hasattr(chart.y_axis, 'scaling'):
                        chart.y_axis.scaling.min = None
                        chart.y_axis.scaling.max = None

                    if hasattr(chart.y_axis, 'majorUnit'):
                        chart.y_axis.majorUnit = None

                    chart.x_axis.axPos = "b"
                    chart.y_axis.axPos = "l"
                    chart.x_axis.tickLblPos = "low"
                    chart.y_axis.tickLblPos = "low"
                    if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
                        chart.x_axis.title.tx.rich.p[0].r.rPr.sz = 1200
                        if hasattr(chart.x_axis.title, 'layout'):
                            chart.x_axis.title.layout = Layout(
                                manualLayout=ManualLayout(
                                    xMode="edge", yMode="edge",
                                    x=0.5, y=0.95, w=0.3, h=0.05
                                )
                            )

                    if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
                        chart.y_axis.title.tx.rich.p[0].r.rPr.sz = 1200

                except AttributeError:
                    pass

                chart_col = current_col + len(df_table.columns) + 1
                chart_col_letter = get_column_letter(chart_col)
                ws.add_chart(chart, f"{chart_col_letter}2")
                
                data_width = len(df_table.columns)
                chart_width = int(chart.width) if hasattr(chart, 'width') else 22

                current_col += data_width + chart_width

            except Exception as e:
                print(f"[PIGNAT ERROR] Failed to process metric: {str(e)}", file=sys.stderr)
                continue
        
        return wb


if __name__ == "__main__":
    try:
        test_dir = "/home/lucaslhm/Bureau/test"
        print(f"📁 Chargement des données depuis: {test_dir}")
        pignat = PignatData(test_dir)
        
        print(f"📊 Données chargées: {len(pignat.data_frame)} lignes")
        print(f"📋 Colonnes disponibles: {pignat.columns}")
        print(f"❌ Colonnes manquantes: {pignat.missing_columns}")
        print(f"✅ Toutes données requises: {pignat.is_all_required_data()}")
        
        print("\n🎯 Test des graphiques disponibles:")
        graphs = pignat.get_available_graphs()
        for graph in graphs:
            status = "✅" if graph['available'] else "❌"
            print(f"   {status} {graph['name']}")
        
        print("\n⏰ Plage temporelle:")
        time_range = pignat.get_time_range()
        print(f"   Min: {time_range['min_time']}")
        print(f"   Max: {time_range['max_time']}")
        print(f"   Points: {len(time_range['unique_times'])}")
        
        print("\n📈 Test d'extraction des métriques:")
        metrics_to_test = [
            TEMPERATURE_DEPENDING_TIME,
            DEBIMETRIC_RESPONSE_DEPENDING_TIME, 
            PRESSURE_PYROLYSEUR_DEPENDING_TIME,
            PRESSURE_POMPE_DEPENDING_TIME,
            DELTA_PRESSURE_DEPENDING_TIME
        ]
        
        for metric in metrics_to_test:
            try:
                data = pignat.get_json_metrics(metric)
                print(f"   ✅ {metric}")
                print(f"      📊 Données: {len(data['data'])} lignes")
                print(f"      🔤 Colonnes: {list(data['data'].columns)}")
                print(f"      📍 Premier échantillon: {data['data'].iloc[0].to_dict()}")
            except Exception as e:
                print(f"   ❌ {metric}: {e}")
        
        print(f"\n📊 Données manquantes par colonne:")
        missing_per_col = pignat.report_missing_per_column()
        for col, count in missing_per_col.items():
            if count > 0:
                print(f"   {col}: {count} valeurs manquantes")
        
        print("\n📝 Génération Excel...")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        metrics_wanted = []
        for graph in graphs:
            if graph['available']:
                metrics_wanted.append({
                    "name": graph['name'],
                    "timeRange": {
                        "startTime": None,
                        "endTime": None
                    }
                })
        
        print(f"📋 Métriques à inclure dans l'Excel: {[m['name'] for m in metrics_wanted]}")
        
        wb = pignat.generate_workbook_with_charts(wb, metrics_wanted)
        wb.save("/home/lucaslhm/Bureau/test_pignat_output.xlsx")
        print("✅ Excel généré: /home/lucaslhm/Bureau/test_pignat_output.xlsx")
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        print(f"📍 Détail: {traceback.format_exc()}")