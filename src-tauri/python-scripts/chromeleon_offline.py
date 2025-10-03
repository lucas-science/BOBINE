from typing import Tuple
import re
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional, Dict, Any, Tuple

MASSE_INJECTEE="masse injectée (kg)"
MASSE_RECETTE="masse recette 1 (kg)"
MASSE_RECETTE2="masse recette 2 (kg)"
MASSE_CENDRIER="masse cendrier (kg)"


class ChromeleonOffline:

    def __init__(self, dir_root: str):
        if not os.path.isdir(dir_root):
            raise FileNotFoundError(f"Le répertoire {dir_root} n'existe pas")

        files = [
            f for f in os.listdir(dir_root)
            if (
                os.path.isfile(os.path.join(dir_root, f))
                and not f.startswith(('.', '~', '.~lock'))
                and f.lower().endswith('.xlsx')
            )
        ]
        if not files:
            raise FileNotFoundError(f"Aucun fichier .xlsx trouvé dans {dir_root}")

        found = {}      
        errors = []

        for fname in files:
            path = os.path.join(dir_root, fname)
            try:
                df = pd.read_excel(
                    path,
                    sheet_name="Integration",
                    header=None,
                    dtype=str
                )
            except Exception as e:
                errors.append(f"{fname}: feuille 'Integration' illisible ({e})")
                continue

            inj_name, tag = self._extract_run_tag_from_df(df)
            if tag is None:
                errors.append(f"{fname}: 'Injection Name' introuvable ou ne contient pas R1/R2")
                continue

            if tag in found:
                prev = found[tag]["file"]
                raise RuntimeError(
                    f"Deux fichiers identifiés comme {tag} :\n"
                    f" - {prev}\n - {path}\n"
                    "Il ne doit y en avoir qu'un seul par série."
                )

            found[tag] = {"df": df, "file": path, "inj_name": inj_name}

        missing = [t for t in ("R1", "R2") if t not in found]
        if missing:
            extra = ("\nDétails :\n- " + "\n- ".join(errors)) if errors else ""
            raise FileNotFoundError(
                f"Impossible d'initialiser : manque {', '.join(missing)}.{extra}"
            )

        self.df_r1 = found["R1"]["df"]
        self.df_r2 = found["R2"]["df"]
        self.file_r1 = found["R1"]["file"]
        self.file_r2 = found["R2"]["file"]
        self.injection_name_r1 = found["R1"]["inj_name"]
        self.injection_name_r2 = found["R2"]["inj_name"]

    def _extract_run_tag_from_df(self, df: pd.DataFrame):
        def is_injection_name(cell) -> bool:
            if not isinstance(cell, str):
                return False
            norm = re.sub(r'[\s:\u00A0]+', '', cell, flags=re.UNICODE).lower()
            return norm == "injectionname"

        nrows, ncols = df.shape
        for i in range(nrows):
            for j in range(ncols):
                if is_injection_name(df.iat[i, j]):
                    value = None
                    for k in range(j + 1, ncols):
                        v = df.iat[i, k]
                        if isinstance(v, str) and v.strip():
                            value = v.strip()
                            break
                        if pd.notna(v) and str(v).strip():
                            value = str(v).strip()
                            break

                    if not value and i + 1 < nrows:
                        v = df.iat[i + 1, j]
                        if isinstance(v, str) and v.strip():
                            value = v.strip()
                        elif pd.notna(v) and str(v).strip():
                            value = str(v).strip()

                    if not value:
                        return None, None
        
                    up = value.upper()
    
                    if re.search(r'\bR1\b', up) or "-R1" in up or "_R1" in up or " R1" in up:
                        return value, "R1"
                    if re.search(r'\bR2\b', up) or "-R2" in up or "_R2" in up or " R2" in up:
                        return value, "R2"
                    return value, None

        return None, None

    def get_graphs_available(self) -> list[dict]:
        graphs = [{
            'name': "Résultats d'intégration de R1 et R2 avec bilan matière",
            'available': False,
        }]
        try:
            if (
                hasattr(self, "df_r1") and not self.df_r1.empty
                and hasattr(self, "df_r2") and not self.df_r2.empty
            ):
                graphs[0]['available'] = True
                return graphs
        except Exception:
            pass

        return []

    def show(self):
        print("Data from R1.xlsx:")
        print(self.df_r1.head())
        print("\nData from R2.xlsx:")
        print(self.df_r2.head())

    def get_R1_R2_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        integration_start_r1 = self.df_r1[self.df_r1[0].str.startswith(
            'Integration Results', na=False)].index[0]
        integration_start_r2 = self.df_r2[self.df_r2[0].str.startswith(
            'Integration Results', na=False)].index[0]

        header_row_r1 = integration_start_r1 + 1
        header_row_r2 = integration_start_r2 + 1
        
        start_r1 = integration_start_r1 + 4
        start_r2 = integration_start_r2 + 4
        
        actual_headers_r1 = []
        actual_headers_r2 = []
        
        for col_idx in range(len(self.df_r1.columns)):
            header_val = self.df_r1.iloc[header_row_r1, col_idx]
            if pd.notna(header_val) and str(header_val).strip():
                actual_headers_r1.append(str(header_val).strip())
            else:
                actual_headers_r1.append(f"Column_{col_idx}")
                
        for col_idx in range(len(self.df_r2.columns)):
            header_val = self.df_r2.iloc[header_row_r2, col_idx]
            if pd.notna(header_val) and str(header_val).strip():
                actual_headers_r2.append(str(header_val).strip())
            else:
                actual_headers_r2.append(f"Column_{col_idx}")

        df1 = self.df_r1.iloc[start_r1:, :].copy()
        df2 = self.df_r2.iloc[start_r2:, :].copy()

        expected_cols = ["No.", "Peakname", "RetentionTime", "Height", "Relative Area", "Amount", "Amount"]
        
        def assign_columns_adaptively(df, actual_headers):
            actual_cols = len(df.columns)
            
            if len(actual_headers) == actual_cols:
                df.columns = actual_headers
            else:
                if actual_cols <= len(expected_cols):
                    df.columns = expected_cols[:actual_cols]
                else:
                    new_cols = expected_cols.copy()
                    for i in range(len(expected_cols), actual_cols):
                        new_cols.append(f"Column_{i+1}")
                    df.columns = new_cols
            
            return df
        
        df1 = assign_columns_adaptively(df1, actual_headers_r1)
        df2 = assign_columns_adaptively(df2, actual_headers_r2)

        def find_column_with_pattern(df, patterns):
            import re
            for col in df.columns:
                col_lower = str(col).lower().strip()
                for pattern, std_name in patterns.items():
                    if re.search(pattern, col_lower, re.IGNORECASE):
                        return col, std_name
            return None, None
        
        patterns = {
            r'^no$|^n°$|^num': 'No.',
            r'peak.*name|peakname': 'Peakname',
            r'ret.*time|retention.*time': 'Retention Time',
            r'^area$|area.*pa.*min': 'Area',
            r'^height$|height.*pa': 'Height',
            r'rel.*area|relative.*area': 'Relative Area',
            r'amount.*%|^amount$': 'Amount (%)',
        }
        
        column_mapping_df1 = {}
        column_mapping_df2 = {}
        
        for col in df1.columns:
            matched_col, std_name = find_column_with_pattern(pd.DataFrame(columns=[col]), patterns)
            if matched_col and std_name:
                column_mapping_df1[col] = std_name
                
        for col in df2.columns:
            matched_col, std_name = find_column_with_pattern(pd.DataFrame(columns=[col]), patterns)
            if matched_col and std_name:
                column_mapping_df2[col] = std_name
        
        df1_final = df1.rename(columns=column_mapping_df1)
        df2_final = df2.rename(columns=column_mapping_df2)
        
        required_cols = ['No.', 'Peakname', 'Retention Time', 'Relative Area']
        available_cols_df1 = [col for col in required_cols if col in df1_final.columns]
        available_cols_df2 = [col for col in required_cols if col in df2_final.columns]
        
        if available_cols_df1:
            df1_final = df1_final[available_cols_df1]
        if available_cols_df2:
            df2_final = df2_final[available_cols_df2]

        return df1_final, df2_final

    def get_relative_area_by_carbon_tables(self) -> dict:
        R1_data, R2_data = self.get_R1_R2_data()

        R1_data = R1_data.copy()
        R2_data = R2_data.copy()
        R1_data['Relative Area'] = pd.to_numeric(
            R1_data['Relative Area'], errors='coerce')
        R2_data['Relative Area'] = pd.to_numeric(
            R2_data['Relative Area'], errors='coerce')

        def process_data(data):
            import re

            results = {}

            carbon_ranges = [f'C{i}' for i in range(6, 33)]


            def is_valid_value(value):
                if pd.isna(value):
                    return False
                if isinstance(value, str):
                    value_clean = value.strip()
                    if value_clean == '' or value_clean.lower() == 'nan':
                        return False
                    try:
                        float(value_clean.replace(',', '.'))
                        return True
                    except (ValueError, TypeError):
                        return False
                try:
                    float(value)
                    return value != 0
                except (ValueError, TypeError):
                    return False

            def convert_to_float(value):
                if isinstance(value, str):
                    return float(value.replace(',', '.'))
                return float(value)

            for carbon in carbon_ranges:
                linear_val = 0
                isomers_val = 0

                linear_patterns = [
                    f'^n-{carbon}$',
                    f'^n{carbon}$',
                    f'^{carbon}\\s*linear$',
                ]

                isomers_patterns = [
                    f'^{carbon}\\s*isomers?$',
                    f'^{carbon}\\s*iso$',
                    f'^iso-{carbon}$',
                ]

                for idx, row in data.iterrows():
                    peakname = str(row.get('Peakname', '')).strip()
                    relative_area = row.get('Relative Area', '')

                    if not peakname or peakname.lower() in ['nan', ''] or not is_valid_value(relative_area):
                        continue

                    try:
                        area_value = convert_to_float(relative_area)
                    except (ValueError, TypeError):
                        continue

                    for pattern in linear_patterns:
                        if re.match(pattern, peakname, re.IGNORECASE):
                            linear_val = area_value
                            break

                    for pattern in isomers_patterns:
                        if re.match(pattern, peakname, re.IGNORECASE):
                            isomers_val = area_value
                            break

                results[carbon] = {
                    'Paraffin': linear_val,
                    'Isomers': isomers_val
                }

            btx_values = {'C6': 0, 'C7': 0, 'C8': 0}
            btx_patterns = {
                'C6': [r'^Benzene-C6$', r'^Benzene$', r'^C6.*benzene$'],
                'C7': [r'^Toluene-C7$', r'^Toluene$', r'^C7.*toluene$'],
                'C8': [r'^Xylenes-C8$', r'^Xylenes$', r'^C8.*xylene', r'^Xylene']
            }

            for idx, row in data.iterrows():
                peakname = str(row.get('Peakname', '')).strip()
                relative_area = row.get('Relative Area', '')

                if not peakname or peakname.lower() in ['nan', ''] or not is_valid_value(relative_area):
                    continue

                try:
                    area_value = convert_to_float(relative_area)
                except (ValueError, TypeError):
                    continue

                for carbon_key, patterns in btx_patterns.items():
                    for pattern in patterns:
                        if re.match(pattern, peakname, re.IGNORECASE):
                            btx_values[carbon_key] = area_value
                            break

            total_linear = sum(results[carbon]['Paraffin'] for carbon in carbon_ranges)
            total_isomers = sum(results[carbon]['Isomers'] for carbon in carbon_ranges)
            total_btx = sum(btx_values.values())


            return results, total_linear, total_isomers, btx_values, total_btx

        results_R1, total_linear_R1, total_isomers_R1, btx_values_R1, total_btx_R1 = process_data(
            R1_data)
        results_R2, total_linear_R2, total_isomers_R2, btx_values_R2, total_btx_R2 = process_data(
            R2_data)

        carbon_ranges = [f'C{i}' for i in range(6, 33)]

        def create_dataframe(results, btx_values, name):
            data_list = []

            for carbon in carbon_ranges:
                linear = results[carbon]['Paraffin']
                isomers = results[carbon]['Isomers']
                btx = btx_values.get(carbon, 0)
                total = linear + isomers + btx

                data_list.append({
                    'Carbon': carbon,
                    'Paraffin': linear,
                    'Isomers': isomers,
                    'BTX': btx,
                    'Total': total
                })

            return pd.DataFrame(data_list)

        df_R1 = create_dataframe(results_R1, btx_values_R1, 'R1')
        df_R2 = create_dataframe(results_R2, btx_values_R2, 'R2')

        df_Moyenne = pd.DataFrame({
            'Carbon': carbon_ranges,
            'Paraffin': [(results_R1[carbon]['Paraffin'] + results_R2[carbon]['Paraffin']) / 2 for carbon in carbon_ranges],
            'Isomers': [(results_R1[carbon]['Isomers'] + results_R2[carbon]['Isomers']) / 2 for carbon in carbon_ranges],
            'BTX': [(btx_values_R1.get(carbon, 0) + btx_values_R2.get(carbon, 0)) / 2 for carbon in carbon_ranges],
            'Total': [((results_R1[carbon]['Paraffin'] + results_R1[carbon]['Isomers'] + btx_values_R1.get(carbon, 0)) +
                      (results_R2[carbon]['Paraffin'] + results_R2[carbon]['Isomers'] + btx_values_R2.get(carbon, 0))) / 2
                      for carbon in carbon_ranges]
        })

        total_identified_R1 = df_R1['Total'].sum()
        total_identified_R2 = df_R2['Total'].sum()
        total_identified_Moyenne = df_Moyenne['Total'].sum()

        autres_R1 = 100 - total_identified_R1
        autres_R2 = 100 - total_identified_R2
        autres_Moyenne = 100 - total_identified_Moyenne

        def add_totals(df, autres_val, total_linear, total_isomers, total_btx):
            totals = pd.DataFrame({
                'Carbon': ['Autres', 'Total'],
                'Paraffin': [0, total_linear],
                'Isomers': [0, total_isomers],
                'BTX': [0, total_btx],
                'Total': [autres_val, 100]
            })
            return pd.concat([df, totals], ignore_index=True)

        df_R1 = add_totals(df_R1, autres_R1, total_linear_R1,
                           total_isomers_R1, total_btx_R1)
        df_R2 = add_totals(df_R2, autres_R2, total_linear_R2,
                           total_isomers_R2, total_btx_R2)
        df_Moyenne = add_totals(df_Moyenne, autres_Moyenne,
                                (total_linear_R1 + total_linear_R2) / 2,
                                (total_isomers_R1 + total_isomers_R2) / 2,
                                (total_btx_R1 + total_btx_R2) / 2)

        return {
            'R1': df_R1,
            'R2': df_R2,
            'Moyenne': df_Moyenne
        }
    

    def _write_df_block(
        self,
        ws: Worksheet,
        title: str,
        df,
        start_col: int,
        start_row: int
    ) -> Tuple[int, int]:
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        gray_fill = PatternFill("solid", fgColor="DDDDDD")
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["No.", "Peakname", "RetentionTime", "Relative Area"]
        subheaders = ["", "", "min", "%"]
        ncols = len(headers)

        ws.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=start_row, end_column=start_col + ncols - 1
        )
        c = ws.cell(row=start_row, column=start_col, value=title)
        c.font = title_font
        c.alignment = Alignment(horizontal="left", vertical="center")

        hr = start_row + 1
        sr = start_row + 2
        for i, h in enumerate(headers):
            col = start_col + i
            ws.cell(row=hr, column=col, value=h).font = header_font
            ws.cell(row=hr, column=col).fill = gray_fill
            ws.cell(row=hr, column=col).alignment = center
            ws.cell(row=hr, column=col).border = border

            ws.cell(row=sr, column=col,
                    value=subheaders[i]).font = header_font if subheaders[i] else Font()
            ws.cell(
                row=sr, column=col).fill = gray_fill if subheaders[i] else PatternFill()
            ws.cell(row=sr, column=col).alignment = center if subheaders[i] else Alignment(
                horizontal="left", vertical="center")
            ws.cell(row=sr, column=col).border = border

        r = sr + 1
        gray_fill = PatternFill("solid", fgColor="DDDDDD")
        header_font = Font(bold=True)
        
        for _, row in df.iterrows():
            peakname = str(row.get("Peakname", "")).strip()
            is_total_row = "Total" in peakname or peakname.lower() in ["total", "total:", "autres"]
            
            no_cell = ws.cell(row=r, column=start_col + 0, value=row["No."])
            no_cell.alignment = right
            no_cell.border = border
            if is_total_row:
                no_cell.fill = gray_fill
            
            peakname_cell = ws.cell(row=r, column=start_col + 1, value=row["Peakname"])
            peakname_cell.border = border
            if is_total_row:
                peakname_cell.fill = gray_fill
                
            try:
                rt = float(str(row["Retention Time"]).replace(",", "."))
            except Exception:
                rt = None
            rt_cell = ws.cell(row=r, column=start_col + 2, value=rt)
            rt_cell.number_format = "0.000"
            rt_cell.border = border
            if is_total_row:
                rt_cell.fill = gray_fill
                
            try:
                ra = float(str(row["Relative Area"]).replace(",", "."))
            except Exception:
                ra = None
            ra_cell = ws.cell(row=r, column=start_col + 3, value=ra)
            ra_cell.number_format = "0.00"
            ra_cell.border = border
            if is_total_row:
                ra_cell.fill = gray_fill
                
            r += 1

        widths = [9, 35, 17, 19]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w

        return r - 1, start_col + ncols - 1

    def _write_bilan_matiere(
        self,
        ws: Worksheet,
        start_col: int,
        start_row: int,
        data: Optional[Dict[str, Any]] = None
    ) -> Tuple[int, int]:
        if not data:
            ws.cell(row=start_row, column=start_col,
                    value="Bilan matière").font = Font(bold=True, size=12)
            return start_row, start_col + 4

        def fmt2(x):
            try:
                return f"{float(x):.2f}".replace(".", ",")
            except Exception:
                return ""

        wt_vals = data.get("wt% R1/R2", [0, 0])
        wt_r1 = fmt2(wt_vals[0]) if len(wt_vals) > 0 else ""
        wt_r2 = fmt2(wt_vals[1]) if len(wt_vals) > 1 else ""
        
        rend = data.get("Rendement (massique)", {})
        
        table_data = []
        table_data.append(["Bilan matière", "", "", "", ""])
        table_data.append(["", "wt% R1/R2", "", "Rendement (massique)", ""])
        table_data.append(["Masse recette 1 (kg)", fmt2(data.get("Masse recette 1 (kg)", 0)), wt_r1, "Liquide (%)", fmt2(rend.get("Liquide (%)", 0))])
        table_data.append(["Masse recette 2 (kg)", fmt2(data.get("Masse recette 2 (kg)", 0)), wt_r2, "Gas (%)", fmt2(rend.get("Gas (%)", 0))])
        table_data.append(["Masse cendrier (kg)", fmt2(data.get("Masse cendrier (kg)", 0)), "", "Residue (%)", fmt2(rend.get("Residue (%)", 0))])
        table_data.append(["Masse injectée (kg)", fmt2(data.get("Masse injectée (kg)", 0)), "", "", ""])

        for row_idx, row_data in enumerate(table_data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

        end_row = start_row + len(table_data) - 1
        end_col = start_col + 4  # 5 colonnes (0-4)
        
        self._apply_bilan_matiere_formatting(ws, start_row, start_col, end_row, end_col)
        
        return end_row, end_col

    def _apply_bilan_matiere_formatting(self, worksheet, start_row, start_col, end_row, end_col):
        thin = Side(style="thin", color="000000")
        thick = Side(style="thick", color="000000")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")

        worksheet.merge_cells(start_row=start_row, start_column=start_col,
                            end_row=start_row, end_column=end_col)
        t = worksheet.cell(row=start_row, column=start_col, value="Bilan matière")
        t.alignment = Alignment(horizontal="center", vertical="center")
        t.font = Font(bold=True)

        header_r = start_row + 1
        cWt1 = start_col + 1
        cWt2 = start_col + 2
        cRL = start_col + 3
        cRV = start_col + 4

        worksheet.cell(row=header_r, column=cWt1, value="wt% R1/R2")
        worksheet.merge_cells(start_row=header_r, start_column=cWt1,
                            end_row=header_r, end_column=cWt2)
        worksheet.cell(row=header_r, column=cWt1).alignment = Alignment(horizontal="center", vertical="center")

        worksheet.cell(row=header_r, column=cRL, value="Rendement (massique)")
        worksheet.merge_cells(start_row=header_r, start_column=cRL,
                            end_row=header_r, end_column=cRV)
        worksheet.cell(row=header_r, column=cRL).alignment = Alignment(horizontal="center", vertical="center")

        for r in range(start_row + 2, end_row + 1):
            worksheet.cell(row=r, column=start_col).alignment = Alignment(horizontal="left", vertical="center")
            worksheet.cell(row=r, column=start_col + 1).alignment = Alignment(horizontal="right", vertical="center")
            worksheet.cell(row=r, column=cWt2).alignment = Alignment(horizontal="right", vertical="center")
            worksheet.cell(row=r, column=cRL).alignment = Alignment(horizontal="left", vertical="center")
            worksheet.cell(row=r, column=cRV).alignment = Alignment(horizontal="right", vertical="center")

        for rr in range(start_row + 2, start_row + 6):
            if start_row <= rr <= end_row:
                worksheet.cell(row=rr, column=start_col + 1).fill = yellow_fill

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                top = thick if r == start_row else thin
                bottom = thick if r == end_row else thin
                left = thick if c == start_col else thin
                right = thick if c == end_col else thin
                worksheet.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)

        worksheet.column_dimensions[get_column_letter(start_col)].width = 25
        worksheet.column_dimensions[get_column_letter(start_col + 1)].width = 15
        worksheet.column_dimensions[get_column_letter(start_col + 2)].width = 12
        worksheet.column_dimensions[get_column_letter(start_col + 3)].width = 20
        worksheet.column_dimensions[get_column_letter(start_col + 4)].width = 15

    @staticmethod
    def compute_bilan(
        masse_injectee: float,
        masse_recette_1: float,
        masse_recette_2: float,
        masse_cendrier: float,
    ) -> Dict[str, Any]:
        if masse_injectee <= 0:
            raise ValueError("masse_injectee doit être > 0")

        m_liquide = masse_recette_1 + masse_recette_2
        m_residu = masse_cendrier
        m_gas = masse_injectee - (m_liquide + m_residu)

        if m_gas < 0:
            m_gas = 0.0

        p_liquide = round(100.0 * m_liquide / masse_injectee, 2)
        p_gas = round(100.0 * m_gas / masse_injectee, 2)
        p_residu = round(100.0 * m_residu / masse_injectee, 2)

        wt_r1 = round(masse_recette_1 / m_liquide, 2) if m_liquide > 0 else 0.0
        wt_r2 = round(masse_recette_2 / m_liquide, 2) if m_liquide > 0 else 0.0

        return {
            "wt% R1/R2": [wt_r1, wt_r2],
            "Masse recette 1 (kg)": masse_recette_1,
            "Masse recette 2 (kg)": masse_recette_2,
            "Masse cendrier (kg)": masse_cendrier,
            "Masse injectée (kg)": masse_injectee,
            "Rendement (massique)": {
                "Liquide (%)": p_liquide,
                "Gas (%)": p_gas,
                "Residue (%)": p_residu,
            },
        }

    def generate_workbook_with_charts(
        self,
        wb: Workbook,
        metrics_wanted: list[str],
        masses: Dict[str, float],
        sheet_name: str = "GC-Offline",
    ) -> Workbook:
        if not metrics_wanted:
            return wb
        else:
            
            bilan_matiere = self.compute_bilan(
                masse_injectee=masses[MASSE_INJECTEE],
                masse_recette_1=masses[MASSE_RECETTE],
                masse_recette_2=masses[MASSE_RECETTE2],
                masse_cendrier=masses[MASSE_CENDRIER],
            )

            ws = wb.create_sheet(title=sheet_name[:31])

            df_r1, df_r2 = self.get_R1_R2_data()

            r1_title = "Liquid Phase Integration Results R1"
            r1_last_row, _ = self._write_df_block(
                ws, r1_title, df_r1, start_col=1, start_row=1)

            r2_title = "Liquid Phase Integration Results R2"
            r2_last_row, _ = self._write_df_block(
                ws, r2_title, df_r2, start_col=9, start_row=1)

            _ = self._write_bilan_matiere(
                ws, start_col=14, start_row=5, data=bilan_matiere)

            start_row = max(r1_last_row, r2_last_row) + 4
            tables = self.get_relative_area_by_carbon_tables()
            
            def write_summary(df, anchor_col, title):
                title_font = Font(bold=True, size=12)
                header_font = Font(bold=True)
                gray_fill = PatternFill("solid", fgColor="DDDDDD")
                center = Alignment(horizontal="center", vertical="center")
                thin = Side(style="thin", color="999999")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                ws.cell(row=start_row, column=anchor_col,
                        value=title).font = title_font

                headers = ["Carbon", "Paraffin", "Isomers", "BTX", "Total"]
                for i, h in enumerate(headers):
                    rr, cc = start_row + 1, anchor_col + i
                    ws.cell(row=rr, column=cc, value=h).font = header_font
                    ws.cell(row=rr, column=cc).fill = gray_fill
                    ws.cell(row=rr, column=cc).alignment = center
                    ws.cell(row=rr, column=cc).border = border

                r = start_row + 2
                for _, row in df.iterrows():
                    carbon_value = row["Carbon"]
                    is_total_row = carbon_value in ["Autres", "Total", "Total:"]
                    
                    # Cellule Carbon
                    carbon_cell = ws.cell(row=r, column=anchor_col + 0, value=carbon_value)
                    carbon_cell.border = border
                    if is_total_row:
                        carbon_cell.fill = gray_fill
                        carbon_cell.font = header_font
                    
                    # Cellules de valeurs
                    for j, key in enumerate(["Paraffin", "Isomers", "BTX", "Total"], start=1):
                        val = float(row[key]) if pd.notna(row[key]) else None
                        c = ws.cell(row=r, column=anchor_col + j, value=val)
                        c.number_format = "0.00"
                        c.border = border
                        if is_total_row:
                            c.fill = gray_fill
                            c.font = header_font
                    r += 1

                widths = [10, 13, 13, 11, 15]
                for i, w in enumerate(widths):
                    ws.column_dimensions[get_column_letter(
                        anchor_col + i)].width = w
                return r

            _ = write_summary(tables["R1"], anchor_col=1,  title="R1")
            _ = write_summary(tables["R2"], anchor_col=9,  title="R2")
            _ = write_summary(tables["Moyenne"], anchor_col=18, title="Moyenne")

            ws.freeze_panes = "A4"
            return wb


# Code de test pour la classe
if __name__ == "__main__":
    try:
        # Test avec le répertoire de données
        print("=== Test de ChromeleonOffline ===")
        off = ChromeleonOffline("/home/lucaslhm/Bureau/test")
        print("✓ Chargement des fichiers R1 et R2 réussi")
        
        # Test 1: Affichage des données brutes
        print("\n=== 1. Données brutes ===")
        off.show()
        
        # Test 2: Données R1 et R2 traitées
        print("\n=== 2. Données R1 et R2 traitées ===")
        df_r1, df_r2 = off.get_R1_R2_data()
        print("DataFrame R1:")
        print(df_r1)
        print("\nDataFrame R2:")
        print(df_r2)
        
        # Test 3: Tableaux de résumé par carbone
        print("\n=== 3. Tableaux de résumé par carbone ===")
        tables = off.get_relative_area_by_carbon_tables()
        
        print("\nTableau R1:")
        print(tables['R1'])
        print("\nTableau R2:")
        print(tables['R2'])
        print("\nTableau Moyenne:")
        print(tables['Moyenne'])
        
        # Test 4: Bilan matière
        print("\n=== 4. Test du bilan matière ===")
        bilan = ChromeleonOffline.compute_bilan(
            masse_injectee=8.0,
            masse_recette_1=1.21,
            masse_recette_2=1.04,
            masse_cendrier=0.59
        )
        print("Bilan matière calculé:")
        for key, value in bilan.items():
            print(f"  {key}: {value}")
        
        # Test 5: Génération du fichier Excel
        print("\n=== 5. Génération du fichier Excel ===")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        wb = off.generate_workbook_with_charts(
            wb,
            metrics_wanted=["Résultats d'intégration de R1 et R2 avec bilan matière"],
            sheet_name="GC Off-line",
            masses={
                MASSE_INJECTEE: 8.0,
                MASSE_RECETTE: 1.21,
                MASSE_RECETTE2: 1.04,
                MASSE_CENDRIER: 0.59
            }
        )
        
        output_path = "/home/lucaslhm/Bureau/chromeleon_offline_test.xlsx"
        wb.save(output_path)
        print(f"✓ Fichier Excel généré avec succès: {output_path}")
        
        # Test 6: Graphiques disponibles
        print("\n=== 6. Graphiques disponibles ===")
        graphs = off.get_graphs_available()
        print("Graphiques disponibles:")
        for graph in graphs:
            status = "✓" if graph['available'] else "✗"
            print(f"  {status} {graph['name']}")
            
        print("\n=== Tous les tests terminés avec succès ===")
        
    except Exception as e:
        print(f"❌ Erreur lors des tests: {e}")
        import traceback
        traceback.print_exc()
