import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Series, Reference
from pigna import PignaData


data = PignaData('/home/lucaslhm/Documents/ETIC/Bobine/project/desktop_app/src-tauri/python-scripts/data/pigna')
metrics_wanted = {
    "chromeleon_online": [],
    "chromeleon_offline": [],
    "pigna": ['Température par rapport au temps','Pression sortie pompe par rapport au temps']
}
dataFromMetricsSensor = {
    "chromeleon_online": [],
    "chromeleon_offline": [],
    "pigna": []
}

def getDataFromMetricsSensor(metrics_wanted: dict[str, list[str]], pignaData):
    if metrics_wanted["chromeleon_offline"] != []:
        # on get les data de chromeleon offline
        pass
    if metrics_wanted["chromeleon_online"] != []:
        # on get les data de chromeleon online
        pass
    print("salut")
    if metrics_wanted["pigna"] != []:
        for metric in metrics_wanted["pigna"]:
            try:
                print("here")
                metric_data = pignaData.get_json_metrics(metric)
                dataFromMetricsSensor['pigna'].append(metric_data)
            except Exception as e:
                print(f"An error occurred: {e}")

    return dataFromMetricsSensor

def save_to_excel_with_charts(data, excel_file_path):
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Créer une feuille pour chaque catégorie
    for category in data:
        if data[category]:  # Vérifier si la catégorie contient des données
            ws = wb.create_sheet(title=category)
            col_offset = 1  # Commencer à la première colonne

            for entry in data[category]:
                df = pd.DataFrame(entry['data'])

                # Écrire les données dans la feuille à partir de la colonne col_offset
                for row in df.itertuples(index=False):
                    ws.append(list(row) if col_offset == 1 else ['']*col_offset + list(row))

                # Créer un graphique pour chaque jeu de données
                chart = LineChart()
                chart.title = entry['name']
                chart.style = 13
                chart.y_axis.title = ', '.join(entry['y_axis']) if len(entry['y_axis'])>1 else entry['y_axis'][0]
                chart.x_axis.title = entry['x_axis']

                data_ref = Reference(ws, min_col=col_offset + 1, min_row=1, max_row=ws.max_row, max_col=col_offset + len(df.columns))
                chart.add_data(data_ref, titles_from_data=True)

                dates_ref = Reference(ws, min_col=col_offset, min_row=2, max_row=ws.max_row)
                chart.set_categories(dates_ref)

                ws.add_chart(chart, f"{chr(69 + col_offset * 3)}{1}")

                col_offset += len(df.columns) + 1  # Ajuster l'offset pour le prochain tableau

    wb.save(excel_file_path)

# Exemple d'utilisation
data = getDataFromMetricsSensor(metrics_wanted, data)
save_to_excel_with_charts(data, 'metrics_data_with_charts.xlsx')