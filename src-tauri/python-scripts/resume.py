import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference, PieChart3D
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList

from typing import Optional, Dict, Any
from chromeleon_online import ChromeleonOnline
from chromeleon_offline import ChromeleonOffline
from context import ExcelContextData



class Resume:
    def __init__(self, dir_online:str, dir_offline:str, dir_context:str):
        """
        Initialize Resume class with ChromeleonOnline, ChromeleonOffline, and ExcelContextData.

        Args:
            dir_root: Root directory containing the data files
        """
        self.dir_online = dir_online
        self.dir_offline = dir_offline
        self.dir_context = dir_context

        self.chromeleon_online = None
        self.chromeleon_offline = None
        self.context_data = None

        # ---- Data ----
        self.masses = {}
        self.online_relative_area_by_carbon = None
        self.offline_relative_area_by_carbon = None

        # Initialize ChromeleonOnline
        try:
            self.chromeleon_online = ChromeleonOnline(dir_online)
            self.online_relative_area_by_carbon = self.chromeleon_online.make_summary_tables()[1]
        except Exception:
            pass

        # Initialize ChromeleonOffline
        try:
            self.chromeleon_offline = ChromeleonOffline(dir_offline)
            self.offline_relative_area_by_carbon = self.chromeleon_offline.get_relative_area_by_carbon_tables()[
                "Moyenne"]
        except Exception:
            pass

        # Initialize ExcelContextData and retrieve masses
        try:
            self.context_data = ExcelContextData(dir_context)
            self.masses = self.context_data.get_masses()
        except Exception:
            pass

    def _get_pourcentage_by_mass(self):
        masse_1 = self.masses.get("masse recette 1 (kg)", 0)
        masse_2 = self.masses.get("masse recette 2 (kg)", 0)
        mass_cendrier = self.masses.get("masse cendrier (kg)", 0)
        mass_total = self.masses.get("masse injectée (kg)", 0)

        liquide_pourcent = (masse_1+masse_2)/mass_total * \
            100 if masse_1 and masse_2 and mass_total else None
        Residue_pourcent = mass_cendrier/mass_total * \
            100 if mass_cendrier and mass_total else None
        Gas_pourcent = 100 - \
            (liquide_pourcent +
             Residue_pourcent) if liquide_pourcent and Residue_pourcent else None

        return {
            "Liquide (%)": liquide_pourcent,
            "Gas (%)": Gas_pourcent,
            "Residue (%)": Residue_pourcent,
        }

    def get_gas_phase(self) -> pd.DataFrame:
        if self.online_relative_area_by_carbon is None:
            return pd.DataFrame()
        
        mass_percentages = self._get_pourcentage_by_mass()
        gas_percent = mass_percentages.get("Gas (%)", 0)
        
        if gas_percent is None or gas_percent == 0:
            return pd.DataFrame()
        
        gas_phase_df = self.online_relative_area_by_carbon.copy()
        
        # Reset index to make Carbon a column instead of index
        gas_phase_df = gas_phase_df.reset_index()
        
        gas_phase_df["% linear"] = gas_phase_df["Linear"] * gas_percent / 100
        gas_phase_df["% iso+olefin"] = gas_phase_df["Olefin"] * gas_percent / 100  
        gas_phase_df["% BTX"] = gas_phase_df["BTX gas"] * gas_percent / 100
        gas_phase_df["% total"] = gas_phase_df["Total"] * gas_percent / 100
        
        return gas_phase_df

    def get_liquid_phase(self) -> pd.DataFrame:
        if self.offline_relative_area_by_carbon is None:
            return pd.DataFrame()
        
        mass_percentages = self._get_pourcentage_by_mass()
        liquide_percent = mass_percentages.get("Liquide (%)", 0)
        
        if liquide_percent is None or liquide_percent == 0:
            return pd.DataFrame()
        
        liquid_phase_df = self.offline_relative_area_by_carbon.copy()
        

        liquid_phase_df = liquid_phase_df.rename(columns={
            'Linear': 'iCn',
            'Isomers': 'nCn'
        })
        
        liquid_phase_df["% iCn"] = liquid_phase_df["iCn"] * liquide_percent / 100
        liquid_phase_df["% nCn"] = liquid_phase_df["nCn"] * liquide_percent / 100
        liquid_phase_df["% BTX"] = liquid_phase_df["BTX"] * liquide_percent / 100
        liquid_phase_df["% Total"] = liquid_phase_df["Total"] * liquide_percent / 100
        
        return liquid_phase_df

    def get_total_phase(self) -> pd.DataFrame:
        gas_phase_df = self.get_gas_phase()
        liquid_phase_df = self.get_liquid_phase()
        
        if gas_phase_df.empty or liquid_phase_df.empty:
            return pd.DataFrame()
        
        carbon_entries = [f'C{i}' for i in range(1, 33)]
        
        result_data = []
        
        for carbon in carbon_entries:
            linear_pct = 0.0
            olefin_pct = 0.0
            btx_pct = 0.0
            total_pct = 0.0
            
            gas_row = gas_phase_df[gas_phase_df['Carbon'] == carbon]
            if not gas_row.empty:
                linear_pct += gas_row['% linear'].iloc[0] if '% linear' in gas_row.columns else 0.0
                olefin_pct += gas_row['% iso+olefin'].iloc[0] if '% iso+olefin' in gas_row.columns else 0.0
                btx_pct += gas_row['% BTX'].iloc[0] if '% BTX' in gas_row.columns else 0.0
                total_pct += gas_row['% total'].iloc[0] if '% total' in gas_row.columns else 0.0
            
            # Get values from liquid phase (if carbon exists in liquid phase)
            liquid_row = liquid_phase_df[liquid_phase_df['Carbon'] == carbon]
            if not liquid_row.empty:
                # For liquid phase: iCn is linear, nCn is olefin
                linear_pct += liquid_row['% iCn'].iloc[0] if '% iCn' in liquid_row.columns else 0.0
                olefin_pct += liquid_row['% nCn'].iloc[0] if '% nCn' in liquid_row.columns else 0.0
                btx_pct += liquid_row['% BTX'].iloc[0] if '% BTX' in liquid_row.columns else 0.0
                total_pct += liquid_row['% Total'].iloc[0] if '% Total' in liquid_row.columns else 0.0
            
            result_data.append({
                'Carbon': carbon,
                '% linear': linear_pct,
                '% olefin': olefin_pct,
                '% BTX': btx_pct,
                '% Total': total_pct
            })
        
        # Add "Autres" and "Total" rows if they exist in either phase
        for special_row in ['Autres', 'Total']:
            linear_pct = 0.0
            olefin_pct = 0.0
            btx_pct = 0.0
            total_pct = 0.0
            
            # Check gas phase
            gas_special = gas_phase_df[gas_phase_df['Carbon'] == special_row]
            if not gas_special.empty:
                linear_pct += gas_special['% linear'].iloc[0] if '% linear' in gas_special.columns else 0.0
                olefin_pct += gas_special['% iso+olefin'].iloc[0] if '% iso+olefin' in gas_special.columns else 0.0
                btx_pct += gas_special['% BTX'].iloc[0] if '% BTX' in gas_special.columns else 0.0
                total_pct += gas_special['% total'].iloc[0] if '% total' in gas_special.columns else 0.0
            
            # Check liquid phase
            liquid_special = liquid_phase_df[liquid_phase_df['Carbon'] == special_row]
            if not liquid_special.empty:
                linear_pct += liquid_special['% iCn'].iloc[0] if '% iCn' in liquid_special.columns else 0.0
                olefin_pct += liquid_special['% nCn'].iloc[0] if '% nCn' in liquid_special.columns else 0.0
                btx_pct += liquid_special['% BTX'].iloc[0] if '% BTX' in liquid_special.columns else 0.0
                total_pct += liquid_special['% Total'].iloc[0] if '% Total' in liquid_special.columns else 0.0
            
            result_data.append({
                'Carbon': special_row,
                '% linear': linear_pct,
                '% olefin': olefin_pct,
                '% BTX': btx_pct,
                '% Total': total_pct
            })
        
        return pd.DataFrame(result_data)


    def get_summary_and_mass_balance(self) -> dict[str, pd.DataFrame]:
        # Get phase tables
        gas_phase_df = self.get_gas_phase()
        liquid_phase_df = self.get_liquid_phase()
        total_phase_df = self.get_total_phase()
        
        if gas_phase_df.empty or liquid_phase_df.empty or total_phase_df.empty:
            return {"summary": pd.DataFrame(), "mass_balance": pd.DataFrame()}
        
        # Helper function to get value from dataframe
        def get_value(df: pd.DataFrame, carbon: str, column: str) -> float:
            """Get value from dataframe for specific carbon and column"""
            row = df[df['Carbon'] == carbon]
            return row[column].iloc[0] if not row.empty and column in row.columns else 0.0
        
        # --- SUMMARY CALCULATIONS ---
        
        # Light olefin = Somme olefin de C2 à C6 dans total phase table
        light_olefin = sum(get_value(total_phase_df, f'C{i}', '% olefin') for i in range(2, 7))
        
        # Aromatics = Somme BTX de C6 à C8 dans total phase table
        aromatics = sum(get_value(total_phase_df, f'C{i}', '% BTX') for i in range(6, 9))
        
        # Other Hydrocarbons gas = %linear de C1 à C8 dans gas phase + C5 et C6 en % iso + olefin dans gas phase + %total de autres dans gas phase table
        other_hc_gas = (
            sum(get_value(gas_phase_df, f'C{i}', '% linear') for i in range(1, 9)) +
            get_value(gas_phase_df, 'C5', '% iso+olefin') +
            get_value(gas_phase_df, 'C6', '% iso+olefin') +
            get_value(gas_phase_df, 'Autres', '% total')
        )
        
        # Other Hydrocarbons liquid = Somme toute valeur de %iCn et %nCn de C6 à C32 dans table liquide phase + %total de autres dans table liquide phase
        other_hc_liquid = (
            sum(get_value(liquid_phase_df, f'C{i}', '% iCn') + get_value(liquid_phase_df, f'C{i}', '% nCn') for i in range(6, 33)) +
            get_value(liquid_phase_df, 'Autres', '% Total')
        )
        
        # Residu = résidu% 
        mass_percentages = self._get_pourcentage_by_mass()
        residue = mass_percentages.get("Residue (%)", 0) or 0
        
        # Individual components from total phase
        ethylene = get_value(total_phase_df, 'C2', '% olefin')
        propylene = get_value(total_phase_df, 'C3', '% olefin') 
        c4_eq = get_value(total_phase_df, 'C4', '% olefin')
        benzene = get_value(total_phase_df, 'C6', '% BTX')
        toluene = get_value(total_phase_df, 'C7', '% BTX')
        xylene = get_value(total_phase_df, 'C8', '% BTX')
        
        # HVC = Ethylène + Propylène + C4 + Benzène + Toluène + Xylène
        hvc = ethylene + propylene + c4_eq + benzene + toluene + xylene
        
        # Create summary DataFrame
        summary_data = {
            "Light olefin": [light_olefin],
            "Aromatics": [aromatics],
            "Other Hydrocarbons gas": [other_hc_gas],
            "Other Hydrocarbons liquid": [other_hc_liquid],
            "Residue": [residue],
            "HVC": [hvc],
            "Ethylene": [ethylene],
            "Propylene": [propylene],
            "C4=": [c4_eq],
            "Benzene": [benzene],
            "Toluene": [toluene],
            "Xylene": [xylene],
        }
        
        summary_df = pd.DataFrame(summary_data)
        
        # --- MASS BALANCE CALCULATIONS ---
        
        # Calculer les wt% R1/R2 en utilisant la même logique que ChromeleonOffline
        masse_recette_1 = self.masses.get("masse recette 1 (kg)", 0)
        masse_recette_2 = self.masses.get("masse recette 2 (kg)", 0)
        m_liquide = masse_recette_1 + masse_recette_2
        
        # Calculer wt% R1/R2 sur la fraction liquide uniquement (comme dans ChromeleonOffline)
        wt_r1 = round(masse_recette_1 / m_liquide, 2) if m_liquide > 0 else 0.0
        wt_r2 = round(masse_recette_2 / m_liquide, 2) if m_liquide > 0 else 0.0

        mass_balance_data = {
            "Flask 1 weight (kg)": [self.masses.get("masse recette 1 (kg)", None)],
            "Flask 2 weight (kg)": [self.masses.get("masse recette 2 (kg)", None)], 
            "Masse cendrier (kg)": [self.masses.get("masse cendrier (kg)", None)],
            "Intrant weight (kg)": [self.masses.get("masse injectée (kg)", None)],
            "wt% R1": [wt_r1],  
            "wt% R2": [wt_r2], 
            "Liquide (%)": [mass_percentages.get("Liquide (%)", None)],
            "Gas (%)": [mass_percentages.get("Gas (%)", None)],
            "Residue (%)": [mass_percentages.get("Residue (%)", None)],
        }
        
        mass_balance_df = pd.DataFrame(mass_balance_data)
        
        return {"summary": summary_df, "mass_balance": mass_balance_df}

    def get_all_graphs_available(self) -> Dict[str, list[dict]]:
        try:
            # Check if we can get summary and mass balance data
            summary_data = self.get_summary_and_mass_balance()
            summary_df = summary_data.get("summary")
            mass_balance_df = summary_data.get("mass_balance")
            
            # Check if we can get total phase data
            total_phase_df = self.get_total_phase()
            
            # Check data availability
            has_summary_data = not summary_df.empty if summary_df is not None else False
            has_mass_balance = not mass_balance_df.empty if mass_balance_df is not None else False
            has_total_phase = not total_phase_df.empty if total_phase_df is not None else False
            
            # Check if we have enough carbons for products repartition
            has_c1_to_c23 = False
            has_c1_to_c8 = False
            
            if has_total_phase:
                carbon_list = total_phase_df['Carbon'].tolist()
                c1_to_c23 = [f'C{i}' for i in range(1, 24)]
                c1_to_c8 = [f'C{i}' for i in range(1, 9)]
                
                has_c1_to_c23 = all(carbon in carbon_list for carbon in c1_to_c23)
                has_c1_to_c8 = all(carbon in carbon_list for carbon in c1_to_c8)
            
            graphs = [
                {
                    'name': "Summary repartition",
                    'available': has_summary_data,
                    'description': "Other Hydrocarbons gas + liquid + Residue + HVC"
                },
                {
                    'name': "HVC Repartition",
                    'available': has_summary_data,
                    'description': "Ethylene + Propylene + C4= + Benzene + Toluene + Xylene"
                },
                {
                    'name': "Phase repartition",
                    'available': has_mass_balance,
                    'description': "Gas%, Liquid%, Residue% distribution"
                },
                {
                    'name': "Products repartition, C1 to C23",
                    'available': has_c1_to_c23,
                    'description': "Total phase composition C1-C23"
                },
                {
                    'name': "Products repartition, C1 to C8",
                    'available': has_c1_to_c8,
                    'description': "Total phase composition C1-C8"
                },
            ]
            
            return graphs
            
        except Exception as e:
            graphs = [
                {
                    'name': "Summary repartition",
                    'available': False
                },
                {
                    'name': "HVC Repartition",
                    'available': False
                },
                {
                    'name': "Phase repartition",
                    'available': False
                },
                {
                    'name': "Products repartition, C1 to C23",
                    'available': False,
                },
                {
                    'name': "Products repartition, C1 to C8",
                    'available': False,
                },
            ]
            
            return graphs


    def _create_summary_repartition_chart(
        self, ws, summary_df, chart_start_row, chart_start_col,
        summary_table_start_col, summary_table_start_row, summary_table_end_row, summary_table_end_col
    ):
        # Colonnes du tableau Summary
        cL2 = summary_table_start_col + 2  # labels centraux
        cV2 = summary_table_start_col + 3  # valeurs centrales

        # Lignes des 4 catégories (dans la zone data ; la 1re ligne data = start_row+1)
        r_gas   = summary_table_start_row + 3  # "Other Hydrocarbons gas"
        r_liq   = summary_table_start_row + 4  # "Other Hydrocarbons liquid"
        r_resid = summary_table_start_row + 5  # "Residue"
        r_hvc   = summary_table_start_row + 6  # "HVC"

        chart = PieChart()
        chart.title = "Summary Repartition"
        data = Reference(ws, min_col=cV2, min_row=r_gas, max_row=r_hvc)
        cats = Reference(ws, min_col=cL2, min_row=r_gas, max_row=r_hvc)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # 1) rendre le camembert bien grand
        chart.height = 15
        chart.width = 24

        # 3) ne pas laisser la légende rétrécir le tracé
        chart.legend.position = "r"   # "t"=top (tu peux garder "r" si tu préfères)
        chart.legend.overlay = True   # la légende “flotte” et n’enlève pas de place au camembert

        # 4) étiquettes propres: Catégorie + %, pas de nom de série ni pastilles carrées
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showCatName   = True
        chart.dataLabels.showPercent   = True
        chart.dataLabels.showVal       = False
        chart.dataLabels.showSerName   = False
        chart.dataLabels.showLegendKey = False   # petits carrés près des étiquettes
        chart.dataLabels.separator     = "; "

        ws.add_chart(chart, ws.cell(row=chart_start_row, column=chart_start_col).coordinate)

    def _create_hvc_repartition_chart(
        self, ws, summary_df, chart_start_row, chart_start_col,
        summary_table_start_col, summary_table_start_row, summary_table_end_row, summary_table_end_col
    ):
        # Paire droite du Summary (labels/valeurs visibles et numériques)
        cL3 = summary_table_start_col + 4
        cV3 = summary_table_start_col + 5

        r1 = summary_table_start_row + 1  # Ethylene
        r6 = summary_table_start_row + 6  # Xylene

        chart = PieChart()
        chart.title = "HVC Repartition"
        data = Reference(ws, min_col=cV3, min_row=r1, max_row=r6)
        cats = Reference(ws, min_col=cL3, min_row=r1, max_row=r6)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = True
        chart.height = 15
        chart.width = 24

        ws.add_chart(chart, ws.cell(row=chart_start_row, column=chart_start_col).coordinate)


    def _create_phase_repartition_chart(
        self, ws, _mass_percentages, chart_start_row, chart_start_col,
        summary_table_start_col, summary_table_start_row
    ):
        # Prend directement les % (numériques) dans la paire gauche du Summary
        cL1 = summary_table_start_col + 0  # labels gauche
        cV1 = summary_table_start_col + 1  # valeurs gauche

        r1 = summary_table_start_row + 1   # %gaz
        r2 = summary_table_start_row + 2   # %liq
        r3 = summary_table_start_row + 3   # % cracking residue

        chart = PieChart3D()
        chart.title = "Phase Repartition"
        data = Reference(ws, min_col=cV1, min_row=r1, max_row=r3)
        cats = Reference(ws, min_col=cL1, min_row=r1, max_row=r3)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = True
        chart.height = 15
        chart.width = 24

        ws.add_chart(chart, ws.cell(row=chart_start_row, column=chart_start_col).coordinate)


    def _create_product_repartition_chart_range(
        self,
        ws,
        total_phase_df: pd.DataFrame,
        chart_start_row: int,
        chart_start_col: int,
        total_table_start_col: int,
        total_table_start_row: int,
        total_table_end_row: int,
        total_table_end_col: int,
        c_start: int,
        c_end: int,
        title: str,
    ):
        """Histogramme empilé (stacked) Linear | Olefin | BTX, plage C{c_start}..C{c_end}."""
        if total_phase_df.empty:
            return

        # Ligne d'en-tête (après le titre du tableau)
        header_row = total_table_start_row + 1
        data_start = header_row + 1            # première ligne data = C1
        data_end   = min(data_start + (c_end - c_start), total_table_end_row)

        # séries = % linear, % olefin, % BTX
        series_min_col = total_table_start_col + 1
        series_max_col = total_table_start_col + 3

        chart = BarChart()
        chart.title = title
        chart.y_axis.title = "Percentage (%)"
        chart.x_axis.title = "Carbon Number"
        chart.type = "col"
        chart.grouping = "stacked"   # <— empilé
        chart.overlap = 100          # empilement visuel franc

        data_ref = Reference(
            ws, min_col=series_min_col, min_row=header_row,
            max_col=series_max_col, max_row=data_end
        )
        cat_ref = Reference(
            ws, min_col=total_table_start_col,
            min_row=data_start, max_row=data_end
        )

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.height = 12
        chart.width = 22

        ws.add_chart(chart, ws.cell(row=chart_start_row, column=chart_start_col).coordinate)



    def generate_workbook_with_charts(
        self,
        wb: Workbook,
        metrics_wanted: list[str],
        sheet_name: str = "Resume",
    ) -> Workbook:
        if not metrics_wanted:
            return wb
            
        try:
            # ---- Build all tables ----
            gas_phase_df = self.get_gas_phase()
            liquid_phase_df = self.get_liquid_phase()
            total_phase_df = self.get_total_phase()
            summary_data = self.get_summary_and_mass_balance()
            summary_df = summary_data["summary"]
            mass_balance_df = summary_data["mass_balance"]

            # Create worksheet
            ws = wb.create_sheet(title=sheet_name[:31])

            # Create two separate tables: Summary and Mass Balance
            def create_summary_table():
                mp = self._get_pourcentage_by_mass() or {}

                def n(v):  # float ou 0.0
                    try:
                        return float(v) if v is not None else 0.0
                    except Exception:
                        return 0.0

                # valeurs numériques depuis summary_df
                get = lambda key: (float(summary_df[key].iloc[0]) if (summary_df is not None and not summary_df.empty) else 0.0)

                data = []
                data.append(["Summary", "", "", "", "", ""])  # titre

                # ligne 1
                data.append(["%gaz", n(mp.get("Gas (%)")),
                            "Light olefin", get("Light olefin"),
                            "Ethylene", get("Ethylene")])

                # ligne 2
                data.append(["%liq", n(mp.get("Liquide (%)")),
                            "Aromatics", get("Aromatics"),
                            "Propylene", get("Propylene")])

                # ligne 3
                data.append(["% cracking residue", n(mp.get("Residue (%)")),
                            "Other Hydrocarbons gas", get("Other Hydrocarbons gas"),
                            "C4=", get("C4=")])

                # ligne 4
                data.append(["", "",
                            "Other Hydrocarbons liquid", get("Other Hydrocarbons liquid"),
                            "Benzene", get("Benzene")])

                # ligne 5
                data.append(["", "",
                            "Residue", get("Residue"),
                            "Toluene", get("Toluene")])

                # ligne 6
                data.append(["", "",
                            "HVC", get("HVC"),
                            "Xylene", get("Xylene")])

                return pd.DataFrame(data)

            def create_mass_balance_table():
                # Get mass percentages
                mp = self._get_pourcentage_by_mass()

                def fmt2(x):
                    try:
                        return f"{float(x):.2f}"
                    except Exception:
                        return ""

                def fmt0(x):
                    try:
                        return f"{float(x):.0f}"
                    except Exception:
                        return ""

                f1 = mass_balance_df['Flask 1 weight (kg)'].iloc[0] if not mass_balance_df.empty else None
                f2w = mass_balance_df['Flask 2 weight (kg)'].iloc[0] if not mass_balance_df.empty else None
                cen = mass_balance_df['Masse cendrier (kg)'].iloc[0] if not mass_balance_df.empty else None
                intr = mass_balance_df['Intrant weight (kg)'].iloc[0] if not mass_balance_df.empty else None

                # 5 colonnes: [Label, R1, R2, Yield-Label, Yield-Value]
                data = []
                data.append(["Mass balance", "", "", "", ""])                               # titre (fusionné)
                data.append(["", "wt% R1/R2", "", "Yield (wieght %)", ""])                  # ligne d'en-tête
                data.append(["Flask 1 weight (kg)", fmt2(f1), "0,54", "Liquide (%)", fmt2(mp.get("Liquide (%)"))])
                data.append(["Flask 2 weight (kg)", fmt2(f2w), "0,46", "Gas (%)",     fmt2(mp.get("Gas (%)"))])
                data.append(["Masse cendrier (kg)", fmt2(cen), "",     "Residue (%)", fmt2(mp.get("Residue (%)"))])
                data.append(["Intrant weight (kg)", fmt0(intr), "",    "",            ""])

                return pd.DataFrame(data)

            
            # Function to apply formatting to Summary table
            def apply_summary_formatting(worksheet, start_row, start_col, end_row, end_col):
                # Define border styles
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Light green fill for specific cells
                light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                # Apply group-specific formatting (skip header row)
                for row_idx in range(1, 7):  # 6 data rows (skip header)
                    current_row = start_row + row_idx
                    
                    # Group 1: Columns 1-2 (%gaz, %liq, % cracking residue) - borders for first 3 rows only
                    if row_idx <= 3:
                        worksheet.cell(row=current_row, column=start_col).border = thin_border
                        worksheet.cell(row=current_row, column=start_col + 1).border = thin_border
                    
                    # Group 2: Columns 3-4 (Light olefin to HVC) - borders + green background
                    worksheet.cell(row=current_row, column=start_col + 2).border = thin_border
                    worksheet.cell(row=current_row, column=start_col + 3).border = thin_border
                    worksheet.cell(row=current_row, column=start_col + 2).fill = light_green_fill
                    worksheet.cell(row=current_row, column=start_col + 3).fill = light_green_fill
                    
                    # Group 3: Columns 5-6 (Ethylene to Xylene) - borders only
                    worksheet.cell(row=current_row, column=start_col + 4).border = thin_border
                    worksheet.cell(row=current_row, column=start_col + 5).border = thin_border
                
                # Apply thick border around entire Summary table (outer perimeter) - AFTER internal borders
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        current_border = cell.border
                        
                        # Keep existing internal borders and add thick outer borders
                        top_border = Side(style='thick') if row == start_row else (current_border.top if current_border else None)
                        bottom_border = Side(style='thick') if row == end_row else (current_border.bottom if current_border else None)
                        left_border = Side(style='thick') if col == start_col else (current_border.left if current_border else None)
                        right_border = Side(style='thick') if col == end_col else (current_border.right if current_border else None)
                        
                        cell.border = Border(
                            top=top_border,
                            bottom=bottom_border, 
                            left=left_border,
                            right=right_border
                        )
                # Affichage "65 %" mais valeur numérique 65.0 (utilisable en chart)
                for rr in range(start_row + 1, start_row + 4):  # %gaz / %liq / % cracking residue
                    ws.cell(row=rr, column=start_col + 1).number_format = '0" %"'  # ou '0.00" %"' si tu veux 2 décimales

                # Valeurs numériques au centre (Light olefin ... HVC)
                for rr in range(start_row + 1, start_row + 7):
                    ws.cell(row=rr, column=start_col + 3).number_format = '0.00'

                # Valeurs numériques à droite (Ethylene ... Xylene)
                for rr in range(start_row + 1, start_row + 7):
                    ws.cell(row=rr, column=start_col + 5).number_format = '0.00'


            def apply_mass_balance_formatting(worksheet, start_row, start_col, end_row, end_col):
                thin  = Side(style="thin",  color="000000")
                thick = Side(style="thick", color="000000")
                yellow_fill = PatternFill("solid", fgColor="FFF2CC")

                # Titre fusionné sur tout le bloc
                worksheet.merge_cells(start_row=start_row, start_column=start_col,
                                    end_row=start_row,   end_column=end_col)
                t = worksheet.cell(row=start_row, column=start_col, value="Mass balance")
                t.alignment = Alignment(horizontal="center", vertical="center")
                t.font = Font(bold=True)

                header_r = start_row + 1
                cR1 = start_col + 1
                cR2 = start_col + 2
                cYL = end_col - 1   # Yield label
                cYV = end_col       # Yield value

                # En-têtes fusionnées
                worksheet.cell(row=header_r, column=cR1, value="wt% R1/R2")
                worksheet.merge_cells(start_row=header_r, start_column=cR1,
                                    end_row=header_r,   end_column=cR2)
                worksheet.cell(row=header_r, column=cR1).alignment = Alignment(horizontal="center", vertical="center")

                worksheet.cell(row=header_r, column=cYL, value="Yield (wieght %)")
                worksheet.merge_cells(start_row=header_r, start_column=cYL,
                                    end_row=header_r,   end_column=cYV)
                worksheet.cell(row=header_r, column=cYL).alignment = Alignment(horizontal="center", vertical="center")

                # Alignements (lignes de données)
                for r in range(start_row + 2, end_row + 1):
                    worksheet.cell(row=r, column=start_col).alignment = Alignment(horizontal="left",  vertical="center")  # labels gauche
                    worksheet.cell(row=r, column=cR1).alignment        = Alignment(horizontal="right", vertical="center") # R1 droite
                    worksheet.cell(row=r, column=cR2).alignment        = Alignment(horizontal="right", vertical="center") # R2 droite
                    worksheet.cell(row=r, column=cYL).alignment        = Alignment(horizontal="left",  vertical="center") # Yield label gauche
                    worksheet.cell(row=r, column=cYV).alignment        = Alignment(horizontal="right", vertical="center") # Yield valeur droite

                # Surlignage jaune pour les cellules saisies (R1 de Flask1/Flask2/Intrant)
                for rr in (start_row + 2, start_row + 3, start_row + 6):
                    if start_row <= rr <= end_row:
                        worksheet.cell(row=rr, column=cR1).fill = yellow_fill

                # Bordures: calcul final par cellule (évite l'écrasement)
                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        top    = thick if r == start_row else thin
                        bottom = thick if r == end_row   else thin
                        left   = thick if c == start_col else thin
                        right  = thick if c == end_col   else thin
                        worksheet.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)


            summary_table = create_summary_table()
            mass_balance_table = create_mass_balance_table()


            # Layout constants
            ROW_GAP = 2      # blank rows between tables vertically
            COL_GAP = 3      # blank columns between blocks horizontally

            def apply_numeric_formatting(worksheet, df, start_row, start_col, title, has_headers=True):
                """
                Apply numeric formatting to display values with 2 decimal places while preserving exact values.
                """
                if df is None or df.empty:
                    return
                    
                # Identify numeric columns (columns with % or known numeric column names)
                numeric_columns = []
                for i, col_name in enumerate(df.columns):
                    if (('%' in str(col_name)) or 
                        any(keyword in str(col_name).lower() for keyword in ['linear', 'olefin', 'btx', 'icn', 'ncn', 'total']) and
                        str(col_name).lower() != 'carbon'):  # Exclude 'Carbon' column
                        numeric_columns.append(i)
                
                # Calculate actual data start row (skip title and headers if they exist)
                data_start_row = start_row + (1 if title else 0) + (1 if has_headers else 0)
                
                # Apply formatting to numeric columns (skip header row)
                for row_idx in range(data_start_row, data_start_row + len(df)):
                    for col_idx in numeric_columns:
                        cell = worksheet.cell(row=row_idx, column=start_col + col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            # Apply number format with 2 decimal places, preserving exact value
                            cell.number_format = '0.00'

            def apply_table_borders(worksheet, start_row, start_col, end_row, end_col, has_title=False):
                """
                Apply logical borders to a table with title.
                - External borders: thick
                - Internal borders: thin
                - Title row gets only external borders if present
                """
                if start_row > end_row or start_col > end_col:
                    return
                    
                thick_border = Side(style="thick", color="000000")
                thin_border = Side(style="thin", color="000000")
                
                # Determine data start row (skip title if present)
                data_start_row = start_row + (1 if has_title else 0)
                
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        
                        # Determine border styles
                        top = thick_border if row == start_row else thin_border
                        bottom = thick_border if row == end_row else thin_border
                        left = thick_border if col == start_col else thin_border  
                        right = thick_border if col == end_col else thin_border
                        
                        # Special case for title row - only external borders
                        if has_title and row == start_row:
                            # Title row: only top, left, right borders (no bottom internal)
                            bottom = None if row < end_row else thick_border
                            
                        # Special case for first data row after title
                        elif has_title and row == data_start_row:
                            # First data row after title gets thick top border
                            top = thick_border
                            
                        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

            def apply_column_width_adjustment(worksheet, df, start_row, start_col, title, has_headers=True):
                """
                Adjust column widths based on formatted content length for a specific table.
                Uses precise width calculation for numeric columns formatted as "0.00"
                """
                if df is None or df.empty:
                    return
                    
                data_start_row = start_row + (1 if title else 0) + (1 if has_headers else 0)
                end_row = data_start_row + len(df) - 1
                
                # Identify numeric columns (same logic as apply_numeric_formatting)
                numeric_columns = set()
                for i, col_name in enumerate(df.columns):
                    if (('%' in str(col_name)) or 
                        any(keyword in str(col_name).lower() for keyword in ['linear', 'olefin', 'btx', 'icn', 'ncn', 'total']) and
                        str(col_name).lower() != 'carbon'):
                        numeric_columns.add(i)
                
                for col_idx in range(len(df.columns)):
                    col_letter = get_column_letter(start_col + col_idx)
                    max_len = 0
                    
                    # Check column header (first row of data which contains column names)
                    if not df.empty:
                        header_value = str(df.columns[col_idx])
                        max_len = max(max_len, len(header_value))
                    
                    # For title length, only consider it for first column since title spans multiple columns
                    if title and col_idx == 0:
                        title_width_per_col = len(title) / len(df.columns)  # Distribute title width
                        max_len = max(max_len, int(title_width_per_col))
                    
                    # Check all data values in this column
                    for row_idx in range(data_start_row, end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=start_col + col_idx)
                        if cell.value is not None:
                            try:
                                if col_idx in numeric_columns and isinstance(cell.value, (int, float)):
                                    # Format as "0.00" to get the actual displayed length
                                    formatted_value = f"{float(cell.value):.2f}"
                                    max_len = max(max_len, len(formatted_value))
                                else:
                                    # Non-numeric values use string representation
                                    value_str = str(cell.value)
                                    max_len = max(max_len, len(value_str))
                            except:
                                continue
                    
                    # Set column width with increased padding for better readability
                    # Comfortable fit for values formatted as "0.00"
                    worksheet.column_dimensions[col_letter].width = min(25, max(6, max_len + 2))

            def apply_wide_column_width_adjustment(worksheet, df, start_row, start_col, title):
                """
                Adjust column widths with generous spacing for Summary and Mass balance tables.
                """
                if df is None or df.empty:
                    return
                    
                data_start_row = start_row + (1 if title else 0)
                end_row = data_start_row + len(df) - 1
                
                for col_idx in range(len(df.columns)):
                    col_letter = get_column_letter(start_col + col_idx)
                    max_len = 0
                    
                    # Check column header if it exists
                    if not df.empty:
                        header_value = str(df.columns[col_idx])
                        max_len = max(max_len, len(header_value))
                    
                    # Check title length for first column
                    if title and col_idx == 0:
                        title_width_per_col = len(title) / len(df.columns)
                        max_len = max(max_len, int(title_width_per_col))
                    
                    # Check all data values in this column
                    for row_idx in range(data_start_row, end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=start_col + col_idx)
                        if cell.value is not None:
                            try:
                                value_str = str(cell.value)
                                max_len = max(max_len, len(value_str))
                            except:
                                continue
                    
                    # Set wider column width with generous padding for readability
                    worksheet.column_dimensions[col_letter].width = min(35, max(12, max_len + 4))

            # Helper to write a dataframe at (start_row, start_col) with a title.
            # Returns (end_row, end_col).
            def write_dataframe(
                df: pd.DataFrame, title: str, start_row: int, start_col: int, include_headers: bool = True
            ) -> tuple[int, int]:
                if df is None or df.empty:
                    return start_row, start_col

                # Calculate number of columns first for title formatting
                ncols = len(df.columns)
                
                # Title (only if not empty) - Bold and centered across table width
                r = start_row
                if title:
                    title_cell = ws.cell(row=start_row, column=start_col, value=title)
                    title_cell.font = Font(bold=True)
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Merge title across table width if table has multiple columns
                    if ncols > 1:
                        end_title_col = start_col + ncols - 1
                        ws.merge_cells(start_row=start_row, start_column=start_col, 
                                     end_row=start_row, end_column=end_title_col)
                    
                    r = start_row + 1  # leave one row between title and header

                # Write dataframe with or without headers
                for row in dataframe_to_rows(df, index=False, header=include_headers):
                    if row is None:
                        continue
                    for c_idx, value in enumerate(row, start=start_col):
                        ws.cell(row=r, column=c_idx, value=value)
                    r += 1

                end_row = r - 1
                end_col = start_col + (ncols - 1 if ncols else 0)
                return end_row, end_col

            # ---------------- Placement ----------------
            current_top_row = 1
            current_left_col = 1

            # 1) Left block: Gas (top-left), Liquid (below)
            gas_end_row, gas_end_col = (current_top_row, current_left_col)
            if not gas_phase_df.empty:
                gas_end_row, gas_end_col = write_dataframe(
                    gas_phase_df, "Gas Phase", current_top_row, current_left_col, include_headers=True
                )
                apply_numeric_formatting(ws, gas_phase_df, current_top_row, current_left_col, "Gas Phase", has_headers=True)
                apply_table_borders(ws, current_top_row, current_left_col, gas_end_row, gas_end_col, has_title=True)
                apply_column_width_adjustment(ws, gas_phase_df, current_top_row, current_left_col, "Gas Phase", has_headers=True)

            liquid_start_row = gas_end_row + ROW_GAP
            liquid_end_row, liquid_end_col = (liquid_start_row, current_left_col)
            if not liquid_phase_df.empty:
                liquid_end_row, liquid_end_col = write_dataframe(
                    liquid_phase_df, "Liquid Phase", liquid_start_row, current_left_col, include_headers=True
                )
                apply_numeric_formatting(ws, liquid_phase_df, liquid_start_row, current_left_col, "Liquid Phase", has_headers=True)
                apply_table_borders(ws, liquid_start_row, current_left_col, liquid_end_row, liquid_end_col, has_title=True)
                apply_column_width_adjustment(ws, liquid_phase_df, liquid_start_row, current_left_col, "Liquid Phase", has_headers=True)

            # The left block width = max of gas/liquid end cols (absolute col index)
            left_block_end_col = max(gas_end_col, liquid_end_col)

            # 2) Middle block: Total phase to the right of both
            total_start_col = max(left_block_end_col, current_left_col) + COL_GAP
            total_end_row, total_end_col = (current_top_row, total_start_col)
            if not total_phase_df.empty:
                total_end_row, total_end_col = write_dataframe(
                    total_phase_df, "Total Phase", current_top_row, total_start_col, include_headers=True
                )
                apply_numeric_formatting(ws, total_phase_df, current_top_row, total_start_col, "Total Phase", has_headers=True)
                apply_table_borders(ws, current_top_row, total_start_col, total_end_row, total_end_col, has_title=True)
                apply_column_width_adjustment(ws, total_phase_df, current_top_row, total_start_col, "Total Phase", has_headers=True)

            # 3) Right block: Summary table and Mass balance table side by side
            summary_start_col = total_end_col + COL_GAP
            summary_end_row, summary_end_col = (current_top_row, summary_start_col)
            if not summary_table.empty:
                summary_end_row, summary_end_col = write_dataframe(
                    summary_table, "", current_top_row, summary_start_col, include_headers=False
                )
                
                # Apply formatting to Summary table
                apply_summary_formatting(ws, current_top_row, summary_start_col, summary_end_row, summary_end_col)
                apply_wide_column_width_adjustment(ws, summary_table, current_top_row, summary_start_col, "")

            # 4) Mass balance table to the right of summary table
            mass_balance_start_col = summary_end_col + COL_GAP
            if not mass_balance_table.empty:
                mb_end_row, mb_end_col = write_dataframe(
                    mass_balance_table, "", current_top_row, mass_balance_start_col, include_headers=False
                )
                apply_mass_balance_formatting(ws, current_top_row, mass_balance_start_col, mb_end_row, mb_end_col)
                apply_wide_column_width_adjustment(ws, mass_balance_table, current_top_row, mass_balance_start_col, "")


            # ---- Add charts based on metrics_wanted ----
            wanted = {m.strip().lower() for m in metrics_wanted}
            def has(name: str) -> bool:
                return name.strip().lower() in wanted

            # Find the bottom of all tables to place charts below
            tables_bottom = max(
                r for r in [
                    locals().get("gas_end_row", 1),
                    locals().get("liquid_end_row", 1),
                    locals().get("total_end_row", 1),
                    locals().get("summary_end_row", 1),
                    locals().get("mb_end_row", 1),
                ] if isinstance(r, int)
            )

            # Dynamic chart positioning - collect charts to create first
            charts_to_create = []
            
            if has("summary repartition"):
                charts_to_create.append(("summary_repartition", "Summary repartition"))
                
            if has("phase repartition"):
                charts_to_create.append(("phase_repartition", "Phase repartition"))
                
            if has("hvc repartition"):
                charts_to_create.append(("hvc_repartition", "HVC repartition"))
                
            if has("products repartition, c1 to c23") or has("products repartition"):
                charts_to_create.append(("products_c1_c23", "Products repartition (C1–C23)"))
                
            if has("products repartition, c1 to c8"):
                charts_to_create.append(("products_c1_c8", "Products repartition (C1–C8)"))

            # Chart layout constants
            CHART_HEIGHT = 35  # Vertical space per chart row
            CHART_WIDTH = 8    # Horizontal space per chart (in columns) - reduced from 14 to 8
            START_ROW = 14     # First chart row
            START_COL = 20     # First chart column (Column T)
            CHARTS_PER_ROW = 2 # Maximum charts per row

            # Calculate dynamic positioning
            current_chart_idx = 0
            for chart_type, _ in charts_to_create:
                # Calculate row and column based on current index
                row_idx = current_chart_idx // CHARTS_PER_ROW
                col_idx = current_chart_idx % CHARTS_PER_ROW
                
                chart_row = START_ROW + (row_idx * CHART_HEIGHT)
                chart_col = START_COL + (col_idx * CHART_WIDTH)
                
                # Create the appropriate chart
                if chart_type == "summary_repartition":
                    self._create_summary_repartition_chart(
                        ws, summary_df, chart_row, chart_col,
                        summary_start_col, current_top_row, summary_end_row, summary_end_col
                    )
                    
                elif chart_type == "phase_repartition":
                    mass_percentages = self._get_pourcentage_by_mass()
                    self._create_phase_repartition_chart(
                        ws, mass_percentages, chart_row, chart_col,
                        summary_start_col, current_top_row
                    )
                    
                elif chart_type == "hvc_repartition":
                    self._create_hvc_repartition_chart(
                        ws, summary_df, chart_row, chart_col,
                        summary_start_col, current_top_row, summary_end_row, summary_end_col
                    )
                    
                elif chart_type == "products_c1_c23":
                    self._create_product_repartition_chart_range(
                        ws, total_phase_df, chart_row, chart_col,
                        total_start_col, current_top_row, total_end_row, total_end_col,
                        c_start=1, c_end=23, title="Products repartition (C1–C23)"
                    )
                    
                elif chart_type == "products_c1_c8":
                    self._create_product_repartition_chart_range(
                        ws, total_phase_df, chart_row, chart_col,
                        total_start_col, current_top_row, total_end_row, total_end_col,
                        c_start=1, c_end=8, title="Products repartition (C1–C8)"
                    )
                
                current_chart_idx += 1

            return wb

        except Exception as e:
            import traceback
            print("generate_workbook_with_charts error:", e)
            traceback.print_exc()
            return wb



if __name__ == "__main__":
    try:
        resume = Resume("/home/lucaslhm/Documents/chromeleon/online","/home/lucaslhm/Documents/chromeleon/offline","/home/lucaslhm/Documents/context/context")
        
        # Test mass percentages
        mass_percentages = resume._get_pourcentage_by_mass()
        print(f"Mass percentages: {mass_percentages}")
        
        # Test gas phase
        gas_phase = resume.get_gas_phase()
        print(f"Gas phase shape: {gas_phase.shape}")
        
        # Test liquid phase
        liquid_phase = resume.get_liquid_phase()
        print(f"Liquid phase shape: {liquid_phase.shape}")
        
        # Debug: Check columns and first few rows
        if not gas_phase.empty:
            print(f"Gas phase columns: {list(gas_phase.columns)}")
            print("Gas phase first 3 rows:")
            print(gas_phase.head(3))
        if not liquid_phase.empty:
            print(f"Liquid phase columns: {list(liquid_phase.columns)}")
        
        # Test total phase
        total_phase = resume.get_total_phase()
        print(f"Total phase shape: {total_phase.shape}")
        
        # Test summary and mass balance
        summary_data = resume.get_summary_and_mass_balance()
        print(f"Summary columns: {list(summary_data['summary'].columns)}")
        print(f"Mass balance columns: {list(summary_data['mass_balance'].columns)}")
        
        # Test graphs availability
        graphs = resume.get_all_graphs_available()
        print("Available graphs:")
        for graph in graphs:
            status = "✓" if graph['available'] else "✗"
            print(f"  {status} {graph['name']}")
        
        # Test new chart-enabled Excel creation
        excel_path = input("Enter Excel output path (or press Enter to skip): ").strip()
        if excel_path:
            # Create workbook
            wb = Workbook()
            
            # Test with all metrics
            all_metrics = [
                "Summary repartition",
                "Phase repartition",
                "Products repartition, C1 to C23",
                "Products repartition, C1 to C8",
                "HVC repartition",
            ]
            
            print(f"Creating Excel with all metrics: {all_metrics}")
            try:
                wb = resume.generate_workbook_with_charts(wb, all_metrics, "Resume_with_Charts")
                
                # Remove default sheet if it exists
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
                
                # Save workbook
                wb.save(excel_path)
                print(f"✓ Excel file with charts created successfully at: {excel_path}")
                print(f"  Sheet created: {wb.sheetnames[0] if wb.sheetnames else 'None'}")
            except Exception as e:
                print(f"✗ Failed to create Excel file: {e}")
                import traceback
                traceback.print_exc()
            
    except Exception as e:
        print("Error in main:", e)

