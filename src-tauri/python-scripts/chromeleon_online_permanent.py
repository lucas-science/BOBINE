import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel

from utils.gc_online.GC_Online_permanent_gas_constants import COMPOUND_MAPPING, CARBON_ROWS, FAMILIES
from utils.time_utils import standardize_injection_time, create_time_sort_key, calculate_total_time_duration
from utils.excel_parsing import find_data_end_row, count_actual_columns, extract_component_blocks, filter_blanc_injections, extract_element_name_adaptive
from utils.excel_formatting import get_standard_styles, get_border, format_table_headers, format_data_table, apply_standard_column_widths, create_title_cell, freeze_panes_standard
from utils.column_mapping import standardize_column_name, get_rel_area_columns, extract_element_names, validate_required_columns
from utils.data_processing import create_summary_table1, create_summary_table2, sort_data_by_time, create_relative_area_summary, process_injection_times, validate_data_availability, calculate_mean_retention_time
from utils.chart_creation import create_chart_configuration, calculate_chart_positions
from utils.file_operations import get_first_excel_file, read_excel_summary, extract_experience_number_adaptive
from utils.chart_styles import apply_line_chart_styles


class ChromeleonOnlinePermanent:
    def __init__(self, dir_root: str, debug: bool = False):
        self.debug = debug
        
        self.first_file = get_first_excel_file(dir_root)
        self.summary_df = read_excel_summary(self.first_file)
        self.experience_number = extract_experience_number_adaptive(self.summary_df)
        
        if isinstance(self.summary_df, pd.DataFrame) and not self.summary_df.empty:
            self.detected_structure = "Summary-based"
        else:
            self.detected_structure = "Unknown"
        
        self.compounds = self._detect_compounds()
    
    
    def _detect_compounds(self):
        compounds = []
        component_blocks = extract_component_blocks(self.summary_df)
        
        for block in component_blocks:
            compound_name = extract_element_name_adaptive(self.summary_df, block['row_index'])
            if compound_name:
                compounds.append({
                    'name': compound_name,
                    'block_start': block['row_index']
                })
        
        return compounds
    
    def get_relative_area_by_injection(self) -> pd.DataFrame:
        data_by_elements = self._extract_compound_data()
        
        if not data_by_elements:
            raise ValueError("Aucun élément chimique trouvé dans les sous-tableaux")

        first_element_df = list(data_by_elements.values())[0]

        required_cols = ['Injection Name', 'Injection Time']
        is_valid, missing = validate_required_columns(first_element_df, required_cols)
        if not is_valid:
            raise ValueError(f"Colonnes manquantes: {missing}. "
                             "Vérifiez que les colonnes 'Inject Time' sont présentes dans les données.")
        
        result = first_element_df[required_cols].copy()
        result = process_injection_times(result)

        for element, df in data_by_elements.items():
            col = f'Rel. Area (%) : {element}'
            if col in df.columns:
                result[col] = pd.to_numeric(df[col], errors='coerce').values

        result = sort_data_by_time(result)

        first_time = str(result['Injection Time'].iloc[0]) if len(result) > 0 else None
        last_time = str(result['Injection Time'].iloc[-1]) if len(result) > 0 else None
        summary = create_relative_area_summary(result, first_time, last_time)

        result = pd.concat([result, pd.DataFrame([summary])], ignore_index=True)
        return result
    
    def _extract_compound_data(self):
        data_by_compound = {}
        
        for comp_info in self.compounds:
            compound_name = comp_info['name']
            block_start = comp_info['block_start']
            
            header_row = block_start + 2
            data_start_row = block_start + 6
            
            if header_row >= len(self.summary_df):
                continue
            
            header_data = self.summary_df.iloc[header_row, :]
            actual_columns = count_actual_columns(header_data)

            if actual_columns == 0:
                header_row = block_start + 3
                header_data = self.summary_df.iloc[header_row, :]
                actual_columns = count_actual_columns(header_data)

            if actual_columns < 6:
                continue
            
            data_end_row = find_data_end_row(self.summary_df, data_start_row, actual_columns)
            temp_df = self.summary_df.iloc[data_start_row:data_end_row, 0:actual_columns].copy()
            temp_df.reset_index(drop=True, inplace=True)
            
            real_headers = header_data.iloc[0:actual_columns].tolist()
            standardized_columns = [standardize_column_name(h, compound_name) for h in real_headers]
            temp_df.columns = standardized_columns
            
            if 'Injection Name' in temp_df.columns:
                temp_df = filter_blanc_injections(temp_df)
                data_by_compound[compound_name] = temp_df
        
        return data_by_compound
    
    def make_summary_tables(self):
        rel_df = self.get_relative_area_by_injection()
        data_by_elements = self._extract_compound_data()
        elements_list = [comp['name'] for comp in self.compounds]
        
        table1 = create_summary_table1(rel_df, data_by_elements, elements_list)
        table2 = create_summary_table2(table1, COMPOUND_MAPPING, CARBON_ROWS, FAMILIES)

        return table1, table2

    def get_graphs_available(self) -> list[dict]:
        graphs = []
        try:
            rel = self.get_relative_area_by_injection()
            validation = validate_data_availability(rel)
            graphs.append({
                'name': "Permanent Gas mass fractions",
                'available': validation['has_enough_timepoints'] and validation['has_numeric_data'],
                'chimicalElements': validation['chemical_elements']
            })
        except Exception:
            graphs.append({'name': "Permanent Gas mass fractions", 'available': False})

        return graphs

    def _calculate_legend_dimensions(self, num_elements: int, legend_position: str = 'b') -> dict:
        """
        Calcule les dimensions optimales pour la légende en fonction du nombre d'éléments.

        Args:
            num_elements: Nombre d'éléments dans la légende
            legend_position: Position de la légende ('b' pour bottom, 'r' pour right, 't' pour top)

        Returns:
            dict avec chart_h, legend_h, legend_y, legend_w, legend_x, chart_height_total
        """
        import math

        if legend_position == 'b':
            # Légende en bas : calculer le nombre de lignes nécessaires
            # Excel affiche environ 3-4 éléments par ligne en légende bottom
            elements_per_row = 4
            num_rows = math.ceil(num_elements / elements_per_row)

            # Chaque ligne nécessite environ 0.035 de hauteur relative
            height_per_row = 0.035
            legend_h = max(0.1, min(0.60, num_rows * height_per_row))  # Entre 10% et 60% pour supporter 50+ éléments

            # CALCUL INVERSÉ : partir du bas pour minimiser le gap
            # Marge en bas pour la légende (2% de padding en bas du graphique Excel)
            bottom_margin = 0.02

            # Position de la légende : la coller en bas avec marge minimale
            legend_y = 1.0 - bottom_margin - legend_h

            # Espace minimal entre plot area et légende (optimisé pour maximiser la zone de tracé)
            axis_and_gap = 0.06  # 6% pour axe X + gap minimal (compact mais sans chevauchement)

            # La plot area se termine juste avant l'axe X
            plot_bottom = legend_y - axis_and_gap

            # La plot area commence après le titre (optimisé pour maximiser la zone de tracé)
            plot_top = 0.06  # 6% en haut (titre compact mais visible)

            # Hauteur de la zone de tracé (MAXIMISÉE pour un graphique bien visible)
            chart_h = max(0.40, plot_bottom - plot_top)  # Minimum 40% pour une zone généreuse

            # Hauteur totale du graphique OPTIMISÉE pour zone de tracé maximale
            if num_elements > 45:
                chart_height_total = 30  # Graphique large pour 45-50 éléments → plot area ~12cm
            elif num_elements > 40:
                chart_height_total = 28  # Graphique large pour 40-45 éléments → plot area ~11cm
            elif num_elements > 35:
                chart_height_total = 25  # Graphique large pour 35-40 éléments → plot area ~10cm
            elif num_elements > 30:
                chart_height_total = 22  # Graphique standard pour 30-35 éléments → plot area ~9cm
            elif num_elements > 20:
                chart_height_total = 20  # Graphique standard pour 20-30 éléments → plot area ~8cm
            else:
                chart_height_total = 15  # Hauteur standard pour <20 éléments → plot area ~6cm

            return {
                'chart_h': chart_h,
                'plot_top': plot_top,  # Position de départ de la plot area (après le titre)
                'legend_h': legend_h,
                'legend_y': legend_y,
                'legend_x': 0.1,
                'legend_w': 0.8,
                'chart_height_total': chart_height_total
            }
        else:
            # Pour les autres positions (right, top), retourner des valeurs par défaut
            return {
                'chart_h': 0.65,
                'plot_top': 0.10,  # Position standard pour autres positions
                'legend_h': 0.1,
                'legend_y': 0.85,
                'legend_x': 0.1,
                'legend_w': 0.8,
                'chart_height_total': 15
            }

    def _calculate_optimal_chart_layout(self, num_elements: int, chart_type: str = "line") -> dict:
        layouts = {
            'line': {
                'mono': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.92, 'h': 0.85},
                    'legend_pos': None,
                    'width': 26, 'height': 15
                },
                'few': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.65, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 26, 'height': 15
                },
                'medium': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.65, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 28, 'height': 15
                },
                'many': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.65, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 32, 'height': 15
                },
                'very_many': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.90, 'h': 0.65},
                    'legend_pos': 'b',
                    'width': 30, 'height': 16
                }
            },
            'bar': {
                'mono': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.90, 'h': 0.85},
                    'legend_pos': None,
                    'width': 22, 'height': 13
                },
                'few': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.75, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 24, 'height': 13
                },
                'medium': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.70, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 26, 'height': 13
                },
                'many': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.60, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 28, 'height': 13
                },
                'very_many': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.85, 'h': 0.70},
                    'legend_pos': 'b',
                    'width': 26, 'height': 13
                }
            }
        }

        if num_elements == 1:
            category = 'mono'
        elif num_elements <= 4:
            category = 'few'
        elif num_elements <= 10:
            category = 'medium'
        elif num_elements <= 20:
            category = 'many'
        else:
            category = 'very_many'

        return layouts[chart_type][category]

    def _apply_ultra_safe_chart_styling(self, chart, chart_type: str = "line"):
        chart.style = 2

        chart.y_axis.title = "Rel. Area (%)" if chart_type == "line" else "Pourcentage (%)"
        chart.x_axis.title = "Injection Time" if chart_type == "line" else "Carbone"

        try:
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            chart.x_axis.crosses = "min"
            chart.y_axis.crosses = "min"
            chart.x_axis.axPos = "b"
            chart.y_axis.axPos = "l"
        except:
            pass

        try:
            chart.y_axis.tickLblPos = "low"
            chart.x_axis.tickLblPos = "low"
        except:
            pass

    def _apply_safe_mono_series_styling(self, chart, num_elements: int):
        try:
            colors = [
                "1f77b4", "ff7f0e", "2ca02c", "d62728", "9467bd", "8c564b",
                "e377c2", "7f7f7f", "bcbd22", "17becf", "aec7e8", "ffbb78",
                "98df8a", "ff9896", "c5b0d5", "c49c94", "f7b6d3", "c7c7c7",
                "dbdb8d", "9edae5", "393b79", "5254a3", "6b6ecf", "9c9ede",
                "ad494a", "8ca252", "bd9e39", "b5cf6b", "cedb9c", "e7ba52",
                "843c39", "de9ed6", "ce6dbd", "756bb1", "9e9ac8", "bcbddc"
            ]

            for i, series in enumerate(chart.series):
                color = colors[i % len(colors)]

                series.smooth = True

                if hasattr(series, 'graphicalProperties'):
                    try:
                        series.graphicalProperties.line.solidFill = color
                        series.graphicalProperties.line.width = 25000
                    except:
                        pass

                try:
                    from openpyxl.chart.marker import Marker
                    series.marker = Marker(symbol="circle", size=5)
                    if hasattr(series.marker, 'graphicalProperties'):
                        series.marker.graphicalProperties.solidFill = color
                except:
                    pass

        except Exception:
            pass

    def generate_workbook_with_charts(self, wb: Workbook, metrics_wanted: list[dict] = None,
                                      sheet_name: str = "GC On-line Permanent Gas") -> Workbook:
        chart_config = create_chart_configuration(metrics_wanted or [])
        asked_names = {(m.get("name") or "").strip() for m in (metrics_wanted or [])}
        chart_config['want_line'] = any(name in asked_names for name in [
            "Permanent Gas mass fractions"
        ])
        
        rel_df = self.get_relative_area_by_injection()
        table1, table2 = self.make_summary_tables()
        ws = wb.create_sheet(title=sheet_name[:31])
        
        styles = get_standard_styles()
        styles['border'] = get_border(styles['black_thin'])
        
        create_title_cell(ws, 1, 1, "%Rel Area par injection (Permanent)", styles)
        headers = list(rel_df.columns)
        start_row = 2
        format_table_headers(ws, headers, start_row, styles=styles)

        ws.row_dimensions[start_row].height = 30

        format_data_table(ws, rel_df, start_row + 1, special_row_identifier="Moyennes", styles=styles)
        apply_standard_column_widths(ws, "main")
        
        table1_row = start_row + len(rel_df) + 3
        create_title_cell(ws, table1_row, 1, "Gas phase Integration Results test average", styles)
        headers1 = ["Peakname", "RetentionTime", "Relative Area"]
        format_table_headers(ws, headers1, table1_row + 1, styles=styles)
        format_data_table(ws, table1, table1_row + 2, special_row_identifier="Total:", styles=styles)
        apply_standard_column_widths(ws, "summary")

        # Adapter la colonne du graphique selon la largeur du tableau rel_df
        from openpyxl.utils import get_column_letter
        num_cols_rel_df = len(rel_df.columns)
        chart_col_index = num_cols_rel_df + 3  # Décalage de 3 colonnes après le tableau
        chart_col = get_column_letter(chart_col_index)
        first_chart_row = 1

        if chart_config['want_line']:
            line_position = f"{chart_col}{first_chart_row}"
            selected_elements = chart_config['selected_elements']

            # Si aucun élément sélectionné, utiliser tous les éléments disponibles
            if not selected_elements:
                all_rel_area_cols = [col for col in headers if col.startswith('Rel. Area (%) :')]
                selected_elements = [col.replace('Rel. Area (%) : ', '') for col in all_rel_area_cols]

            if selected_elements:
                num_elements = len(selected_elements)

                layout_config = self._calculate_optimal_chart_layout(num_elements, "line")

                # Calculer les dimensions optimales de la légende en fonction du nombre d'éléments
                legend_dims = self._calculate_legend_dimensions(num_elements, 'b')

                line_chart = LineChart()
                line_chart.title = "Permanent Gas mass fractions"

                self._apply_ultra_safe_chart_styling(line_chart, "line")

                line_chart.width = layout_config['width']
                line_chart.height = legend_dims['chart_height_total']  # Hauteur ajustée dynamiquement

                try:
                    from openpyxl.chart.layout import Layout, ManualLayout
                    # Ajuster la zone du graphique avec les dimensions calculées dynamiquement
                    line_chart.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode="edge", yMode="edge",
                            x=0.1,   # Marge gauche pour titre Y
                            y=legend_dims['plot_top'],  # Position calculée dynamiquement (après le titre)
                            w=0.75,  # Largeur réduite pour espace légende
                            h=legend_dims['chart_h']  # Hauteur calculée dynamiquement pour maximiser la zone de tracé
                        )
                    )
                except:
                    pass

                if num_elements == 1:
                    line_chart.legend = None
                else:
                    line_chart.legend.position = 'b'  # Bottom position
                    line_chart.legend.overlay = False

                    # Layout manuel avec dimensions calculées dynamiquement pour accommoder tous les éléments
                    try:
                        from openpyxl.chart.layout import Layout, ManualLayout
                        line_chart.legend.layout = Layout(
                            manualLayout=ManualLayout(
                                xMode="edge", yMode="edge",
                                x=legend_dims['legend_x'],
                                y=legend_dims['legend_y'],
                                w=legend_dims['legend_w'],
                                h=legend_dims['legend_h']  # Hauteur calculée pour afficher tous les éléments
                            )
                        )
                    except:
                        line_chart.legend.position = 'b'

                data_df = rel_df[rel_df['Injection Name'] != 'Moyennes'].copy()
                data_rows_count = len(data_df)

                y_cols = []
                for element in selected_elements:
                    col_name = f'Rel. Area (%) : {element}'
                    if col_name in headers:
                        y_cols.append(col_name)

                if y_cols and data_rows_count > 0:
                    # Ajouter chaque série individuellement pour éviter les colonnes intermédiaires non désirées
                    for i, col in enumerate(y_cols):
                        col_index = headers.index(col) + 1
                        data_ref = Reference(ws,
                                           min_col=col_index,
                                           min_row=start_row + 1,  # Données sans header
                                           max_col=col_index,
                                           max_row=start_row + data_rows_count)
                        line_chart.add_data(data_ref, titles_from_data=False)

                        # Définir manuellement le titre de la série
                        element_name = col.replace('Rel. Area (%) : ', '')
                        if i < len(line_chart.series):
                            try:
                                from openpyxl.chart.series import SeriesLabel
                                series_label = SeriesLabel()
                                series_label.v = element_name
                                line_chart.series[i].tx = series_label
                            except:
                                pass

                    time_col_index = headers.index('Injection Time') + 1
                    cats = Reference(ws,
                                   min_col=time_col_index,
                                   min_row=start_row + 1,
                                   max_row=start_row + data_rows_count)
                    line_chart.set_categories(cats)

                self._apply_safe_mono_series_styling(line_chart, num_elements)

                # Appliquer la charte graphique (Futura PT Medium 18 pour titre, légende en bas)
                # preserve_legend_layout=True pour garder le layout dynamique calculé ci-dessus
                apply_line_chart_styles(line_chart, "Permanent Gas mass fractions", legend_position='b', preserve_legend_layout=True)

                ws.add_chart(line_chart, line_position)

        return wb


if __name__ == "__main__":
    import sys
    from datetime import datetime

    if len(sys.argv) > 1:
        data_path = sys.argv[1]
    else:
        data_path = "C:/Users/lucas/Desktop/test"

    try:
        print("[ANALYSE] === CHROMELEON PERMANENT GAS ===")
        print(f"[PATH] Chemin d'analyse: {data_path}")
        
        start_time = datetime.now()
        analyzer = ChromeleonOnlinePermanent(data_path, debug=True)
        init_time = datetime.now() - start_time
        
        print(f"[TIME] Initialisation: {init_time.total_seconds():.2f}s")
        print(f"[EXP] Expérience détectée: {analyzer.experience_number or 'Non définie'}")
        print(f"[STRUCT] Structure détectée: {analyzer.detected_structure}")
        print(f"[COMP] Composés détectés: {len(analyzer.compounds)}")
        

        # Analyse des données extraites
        print("\n[DATA] === ANALYSE DES DONNÉES EXTRAITES ===")
        
        extract_start = datetime.now()
        rel_data = analyzer.get_relative_area_by_injection()
        extract_time = datetime.now() - extract_start
        
        print(f"[TIME] Extraction: {extract_time.total_seconds():.2f}s")
        print(f"[ROWS] Lignes de données: {len(rel_data)}")
        print(f"[COLS] Colonnes de données: {len([c for c in rel_data.columns if c.startswith('Rel. Area')])}")
        
        print("[OK] Extraction terminée")

        # Test graphiques disponibles
        print("\n[CHARTS] === GRAPHIQUES ET MÉTRIQUES DISPONIBLES ===")
        graphs_info = analyzer.get_graphs_available()
        
        available_count = sum(1 for g in graphs_info if g.get('available'))
        print(f"[AVAILABLE] Graphiques disponibles: {available_count}/{len(graphs_info)}")
        
        for g in graphs_info:
            status = "[OK] DISPONIBLE" if g.get('available') else "[NO] INDISPONIBLE"
            print(f"   - {g['name']}: {status}")
            
            if 'chimicalElements' in g and g['chimicalElements']:
                elements_display = ', '.join(g['chimicalElements'][:3])
                if len(g['chimicalElements']) > 3:
                    elements_display += f" et {len(g['chimicalElements'])-3} autres"
                print(f"     Elements: {elements_display}")

        # Construction des métriques pour génération
        metrics = []
        for g in graphs_info:
            if g.get("available"):
                m = {"name": g["name"]}
                if "chimicalElements" in g and g["chimicalElements"]:
                    m["chimicalElementSelected"] = g["chimicalElements"][:5]
                metrics.append(m)

        # Test de génération Excel
        print("\n[EXCEL] === GÉNÉRATION DU RAPPORT EXCEL ===")
        
        gen_start = datetime.now()
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        sheet_name = "GC-Permanent"
        wb = analyzer.generate_workbook_with_charts(
            wb=wb,
            metrics_wanted=metrics,
            sheet_name=sheet_name
        )

        # Sauvegarde
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = analyzer.experience_number or "GC"
        out_name = f"{base}_Permanent_Analysis_{timestamp}.xlsx"
        out_path = os.path.join(data_path, out_name)

        wb.save("C:/Users/lucas/Desktop/test" + out_name)
        gen_time = datetime.now() - gen_start
        
        print(f"[TIME] Génération: {gen_time.total_seconds():.2f}s")
        print(f"[FILE] Fichier généré: {out_path}")

        # Résumé final
        total_time = datetime.now() - start_time
        print(f"\n[SUCCESS] === RÉSUMÉ D'ANALYSE RÉUSSIE ===")
        print(f"[TIME] Temps total: {total_time.total_seconds():.2f}s")
        print(f"[COMP] Composés analysés: {len(analyzer.compounds)}")
        print(f"[ROWS] Lignes de données: {len(rel_data) if len(rel_data) > 0 else 'Aucune'}")
        print(f"[CHARTS] Graphiques générés: {available_count}")
        print(f"[OK] Analyse terminée avec succès")
        

    except Exception as e:
        print(f"\n[ERROR] === ERREUR DURANT L'ANALYSE ===")
        print(f"[ERR] Erreur: {str(e)}")
        print(f"[TYPE] Type: {type(e).__name__}")
        
        # Tentative de diagnostic d'erreur
        try:
            if 'analyzer' in locals():
                print(f"[STATE] État de l'analyseur:")
                print(f"   - Composés détectés: {len(analyzer.compounds) if hasattr(analyzer, 'compounds') else 'N/A'}")
                print(f"   - Structure: {analyzer.detected_structure if hasattr(analyzer, 'detected_structure') else 'N/A'}")
        except:
            pass
        
        import traceback
        print(f"[TRACE] Trace détaillée:")
        traceback.print_exc()
        sys.exit(1)
