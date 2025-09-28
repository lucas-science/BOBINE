"""
ChromeleonOnline - Processeur de données avec graphiques Excel
"""
import os
import re
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference

from utils.gc_online.GC_Online_constants import COMPOUND_MAPPING, CARBON_ROWS, FAMILIES, HVC_CATEGORIES
from utils.time_utils import standardize_injection_time, create_time_sort_key, calculate_total_time_duration
from utils.excel_parsing import find_data_end_row, count_actual_columns, extract_component_blocks, filter_blanc_injections
from utils.excel_formatting import get_standard_styles, get_border, format_table_headers, format_data_table, apply_standard_column_widths, create_title_cell, freeze_panes_standard
from utils.column_mapping import standardize_column_name, get_rel_area_columns, extract_element_names, validate_required_columns
from utils.data_processing import create_summary_table1, process_table1_with_grouping, create_summary_table2, sort_data_by_time, create_relative_area_summary, process_injection_times, validate_data_availability
from utils.chart_creation import create_chart_configuration, calculate_chart_positions
from utils.file_operations import get_first_excel_file, read_excel_summary, extract_experience_number_simple

class ChromeleonOnline:
    def __init__(self, dir_root: str):
        self.first_file = get_first_excel_file(dir_root)
        self.df = read_excel_summary(self.first_file)
        self.experience_number = extract_experience_number_simple(self.df)

    def get_graphs_available(self) -> list[dict]:
        graphs = []

        # Graphique 1: %mass gaz en fonction du temps
        try:
            rel = self.get_relative_area_by_injection()
            validation = validate_data_availability(rel)
            
            graphs.append({
                'name': '%mass gaz en fonction du temps',
                'available': validation['has_enough_timepoints'] and validation['has_numeric_data'],
                'chimicalElements': validation['chemical_elements']
            })
        except Exception:
            graphs.append({
                'name': '%mass gaz en fonction du temps',
                'available': False,
            })

        # Graphique 2: products repartition gaz phase
        try:
            _, table2 = self.make_summary_tables()
            fam_cols = [c for c in ['Linear', 'Olefin', 'BTX gas'] if c in table2.columns]
            has_nonzero = (table2[fam_cols].to_numpy().sum() > 0) if fam_cols else False

            graphs.append({
                'name': 'products repartition gaz phase',
                'available': bool(has_nonzero)
            })
        except Exception:
            graphs.append({
                'name': 'products repartition gaz phase',
                'available': False
            })

        return graphs

    def _get_data_by_elements(self):
        data_by_injection = {}
        
        component_blocks = extract_component_blocks(self.df)

        for block in component_blocks:
            element_name = block['element_name']
            if not element_name:
                continue

            header_row = block['header_row']
            data_start_row = block['data_start_row']
            header_data = self.df.iloc[header_row, :]

            actual_columns = count_actual_columns(header_data)

            if actual_columns == 0:
                header_row = block['row_index'] + 3
                header_data = self.df.iloc[header_row, :]
                actual_columns = count_actual_columns(header_data)

            if actual_columns < 6:
                continue

            data_end_row = find_data_end_row(self.df, data_start_row, actual_columns)
            
            temp_df = self.df.iloc[data_start_row:data_end_row, 0:actual_columns].copy()
            temp_df.reset_index(drop=True, inplace=True)
            
            real_headers = header_data.iloc[0:actual_columns].tolist()
            standardized_columns = []

            for real_header in real_headers:
                standardized_name = standardize_column_name(real_header, element_name)
                standardized_columns.append(standardized_name)

            temp_df.columns = standardized_columns

            if 'Injection Name' not in temp_df.columns:
                continue

            temp_df = filter_blanc_injections(temp_df)
            data_by_injection[element_name] = temp_df

        return data_by_injection

    def get_relative_area_by_injection(self) -> pd.DataFrame:
        data_by_elements = self._get_data_by_elements()
        
        if not data_by_elements:
            raise ValueError("Aucun élément chimique trouvé dans les sous-tableaux")

        first_element_df = list(data_by_elements.values())[0]

        required_cols = ['Injection Name', 'Injection Time']
        is_valid, missing = validate_required_columns(first_element_df, required_cols)
        if not is_valid:
            raise ValueError(f"Colonnes manquantes: {missing}. "
                           "Vérifiez que les colonnes 'Inject Time' sont présentes dans les données.")
        
        result = first_element_df[required_cols].copy()
        
        result = process_injection_times(result)

        for element, df in data_by_elements.items():
            col = f'Rel. Area (%) : {element}'
            if col in df.columns:
                result[col] = pd.to_numeric(df[col], errors='coerce').values

        result = sort_data_by_time(result)

        first_time = str(result['Injection Time'].iloc[0]) if len(result) > 0 else None
        last_time = str(result['Injection Time'].iloc[-1]) if len(result) > 0 else None
        summary = create_relative_area_summary(result, first_time, last_time)

        result = pd.concat([result, pd.DataFrame([summary])], ignore_index=True)
        return result

    def make_summary_tables(self):
        rel_df = self.get_relative_area_by_injection()
        data_by_elements = self._get_data_by_elements()
        
        table1 = create_summary_table1(rel_df, data_by_elements)
        
        table1 = process_table1_with_grouping(table1)
        
        table2 = create_summary_table2(table1, COMPOUND_MAPPING, CARBON_ROWS, FAMILIES)

        return table1, table2

    def _calculate_optimal_chart_layout(self, num_elements: int, chart_type: str = "line") -> dict:
        layouts = {
            'line': {
                'mono': {  # 1 élément - pas de légende
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.92, 'h': 0.85},
                    'legend_pos': None,
                    'width': 26, 'height': 15
                },
                'few': {  # 2-4 éléments - légende droite compacte
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.65, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 26, 'height': 15
                },
                'medium': {  # 5-10 éléments - légende droite élargie
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.65, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 28, 'height': 15
                },
                'many': {  # 11-20 éléments - légende droite très large
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.65, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 32, 'height': 15
                },
                'very_many': {  # 21+ éléments - légende en bas sur plusieurs colonnes
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.90, 'h': 0.65},
                    'legend_pos': 'b',
                    'width': 30, 'height': 16
                }
            },
            'bar': {
                'mono': {
                    'chart': {'x': 0.05, 'y': 0.05, 'w': 0.90, 'h': 0.85},
                    'legend_pos': None,
                    'width': 22, 'height': 13
                },
                'few': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.75, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 24, 'height': 13
                },
                'medium': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.70, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 26, 'height': 13
                },
                'many': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.60, 'h': 0.85},
                    'legend_pos': 'r',
                    'width': 28, 'height': 13
                },
                'very_many': {
                    'chart': {'x': 0.08, 'y': 0.05, 'w': 0.85, 'h': 0.70},
                    'legend_pos': 'b',
                    'width': 26, 'height': 13
                }
            }
        }

        if num_elements == 1:
            category = 'mono'
        elif num_elements <= 4:
            category = 'few'
        elif num_elements <= 10:
            category = 'medium'
        elif num_elements <= 20:
            category = 'many'
        else:
            category = 'very_many'

        return layouts[chart_type][category]

    def _apply_ultra_safe_chart_styling(self, chart, chart_type: str = "line"):
        chart.style = 2

        chart.y_axis.title = "Rel. Area (%)" if chart_type == "line" else "Pourcentage (%)"
        chart.x_axis.title = "Injection Time" if chart_type == "line" else "Carbone"

        try:
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            chart.x_axis.crosses = "min"
            chart.y_axis.crosses = "min"
            chart.x_axis.axPos = "b"
            chart.y_axis.axPos = "l"
        except:
            pass

        try:
            chart.y_axis.tickLblPos = "low"
            chart.x_axis.tickLblPos = "low"
        except:
            pass

    def _apply_ultra_safe_tick_interval(self, chart, num_data_points: int):
        pass

    def _apply_safe_mono_series_styling(self, chart, num_elements: int):
        try:
            colors = [
                "1f77b4", "ff7f0e", "2ca02c", "d62728", "9467bd", "8c564b",
                "e377c2", "7f7f7f", "bcbd22", "17becf", "aec7e8", "ffbb78",
                "98df8a", "ff9896", "c5b0d5", "c49c94", "f7b6d3", "c7c7c7",
                "dbdb8d", "9edae5", "393b79", "5254a3", "6b6ecf", "9c9ede",
                "ad494a", "8ca252", "bd9e39", "b5cf6b", "cedb9c", "e7ba52",
                "843c39", "de9ed6", "ce6dbd", "756bb1", "9e9ac8", "bcbddc"
            ]

            for i, series in enumerate(chart.series):
                color = colors[i % len(colors)]

                series.smooth = True

                if hasattr(series, 'graphicalProperties'):
                    try:
                        series.graphicalProperties.line.solidFill = color
                        series.graphicalProperties.line.width = 25000
                    except:
                        pass

                try:
                    from openpyxl.chart.marker import Marker
                    series.marker = Marker(symbol="circle", size=5)
                    if hasattr(series.marker, 'graphicalProperties'):
                        series.marker.graphicalProperties.solidFill = color
                except:
                    pass

        except Exception:
            pass


    def generate_workbook_with_charts(
        self,
        wb: Workbook,
        metrics_wanted: list[dict],
        sheet_name: str = "GC-Online",
    ) -> Workbook:
        chart_config = create_chart_configuration(metrics_wanted)
        
        rel_df = self.get_relative_area_by_injection()
        table1, table2 = self.make_summary_tables()
        
        ws = wb.create_sheet(title=sheet_name[:31])
        
        styles = get_standard_styles()
        styles['border'] = get_border(styles['black_thin'])
        
        create_title_cell(ws, 1, 1, "%Rel Area par injection (Online)", styles)
        
        headers = list(rel_df.columns)
        start_row = 2
        
        format_table_headers(ws, headers, start_row, styles=styles)
        
        format_data_table(ws, rel_df, start_row + 1, special_row_identifier="Moyennes", styles=styles)
        
        apply_standard_column_widths(ws, "main")
        
        table1_row = start_row + len(rel_df) + 8
        create_title_cell(ws, table1_row, 1, "Gas phase Integration Results test average", styles)
        
        headers1 = ["Peakname", "RetentionTime", "Relative Area"]
        format_table_headers(ws, headers1, table1_row + 1, styles=styles)
        format_data_table(ws, table1, table1_row + 2, special_row_identifier="Total:", styles=styles)
        apply_standard_column_widths(ws, "summary")
        
        table2_col = 6
        table2_row = table1_row
        
        if not table2.empty:
            create_title_cell(ws, table2_row, table2_col, "Regroupement par carbone / famille", styles)
            
            headers2 = ["Carbon"] + list(table2.columns)
            format_table_headers(ws, headers2, table2_row + 1, table2_col, styles=styles)
            
            r = table2_row + 2
            for idx, row in table2.reset_index().iterrows():
                is_total_row = str(row["Carbon"]).lower() == "total"
                for j, colname in enumerate(["Carbon"] + list(table2.columns)):
                    val = row[colname] if colname in row else 0
                    cell = ws.cell(row=r, column=table2_col + j, value=val)
                    cell.border = styles['border']
                    if isinstance(val, (int, float)) and colname != "Carbon":
                        cell.number_format = "0.00"
                    if is_total_row:
                        cell.fill = styles['gray_fill']
                r += 1
            
            apply_standard_column_widths(ws, "carbon_family")
        
        hvc_col = 12
        create_title_cell(ws, table1_row, hvc_col, "composition moyenne principaux HVC (%)", styles)
        
        hvc_headers = ["Molécule", "Moyenne (%)"]
        format_table_headers(ws, hvc_headers, table1_row + 1, hvc_col, styles=styles)
        
        hvc_data = []
        for display_name, carbon, family in HVC_CATEGORIES:
            try:
                if carbon in table2.index and family in table2.columns:
                    val = float(table2.loc[carbon, family])
                else:
                    val = 0.0
            except:
                val = 0.0
            hvc_data.append({"Molécule": display_name, "Moyenne (%)": val})
        
        hvc_df = pd.DataFrame(hvc_data)
        format_data_table(ws, hvc_df, table1_row + 2, hvc_col, styles=styles)
        apply_standard_column_widths(ws, "hvc")
        
        chart_col = "P"
        first_chart_row = table1_row
        
        graphs_to_create = []
        if chart_config['want_line']:
            graphs_to_create.append("line")
        if chart_config['want_bar']:
            graphs_to_create.append("bar")
        
        chart_positions = calculate_chart_positions(graphs_to_create, first_chart_row)

        if len(graphs_to_create) == 2:
            separation_offset = 22  # Plus d'espace entre les graphiques
        else:
            separation_offset = 8

        if chart_config['want_line']:
            line_position = f"{chart_col}{chart_positions.get('line', first_chart_row)}"
            selected_elements = chart_config['selected_elements']

            # Si aucun élément sélectionné, utiliser tous les éléments disponibles
            if not selected_elements:
                all_rel_area_cols = [col for col in headers if col.startswith('Rel. Area (%) :')]
                selected_elements = [col.replace('Rel. Area (%) : ', '') for col in all_rel_area_cols]

            num_elements = len(selected_elements)

            layout_config = self._calculate_optimal_chart_layout(num_elements, "line")

            line_chart = LineChart()
            line_chart.title = "%mass gaz en fonction du temps"

            self._apply_ultra_safe_chart_styling(line_chart, "line")

            line_chart.width = layout_config['width']
            line_chart.height = layout_config['height']

            try:
                from openpyxl.chart.layout import Layout, ManualLayout
                # Ajuster la zone du graphique pour laisser de l'espace aux titres
                line_chart.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode="edge", yMode="edge",
                        x=0.1,   # Marge gauche pour titre Y
                        y=0.1,   # Marge haute pour titre principal
                        w=0.75,  # Largeur réduite pour espace légende
                        h=0.65   # Hauteur réduite pour espace légende en bas
                    )
                )
            except:
                pass

            if num_elements == 1:
                line_chart.legend = None
            else:
                line_chart.legend.position = 'b'  # Bottom position
                line_chart.legend.overlay = False

                # Layout manuel SAFE pour la légende en bas
                try:
                    from openpyxl.chart.layout import Layout, ManualLayout
                    line_chart.legend.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode="edge", yMode="edge",
                            x=0.1,   # Centré horizontalement
                            y=0.85,  # En bas du graphique
                            w=0.8,   # Largeur pour s'étaler
                            h=0.1    # Hauteur compacte
                        )
                    )
                except:
                    line_chart.legend.position = 'b'

            data_df = rel_df[rel_df['Injection Name'] != 'Moyennes'].copy()
            data_rows_count = len(data_df)

            y_cols = []
            for element in selected_elements:
                col_name = f'Rel. Area (%) : {element}'
                if col_name in headers:
                    y_cols.append(col_name)

            if y_cols and data_rows_count > 0:
                y_col_indices = [headers.index(col) + 1 for col in y_cols]

                min_col_y = min(y_col_indices)
                max_col_y = max(y_col_indices)
                data_ref = Reference(ws,
                                   min_col=min_col_y,
                                   min_row=start_row,  # Inclut header
                                   max_col=max_col_y,
                                   max_row=start_row + data_rows_count)
                line_chart.add_data(data_ref, titles_from_data=True)

                time_col_index = headers.index('Injection Time') + 1
                cats = Reference(ws,
                               min_col=time_col_index,
                               min_row=start_row + 1,
                               max_row=start_row + data_rows_count)
                line_chart.set_categories(cats)

                try:
                    from openpyxl.chart.series import SeriesLabel
                    for i, series in enumerate(line_chart.series):
                        if i < len(selected_elements):
                            element_name = selected_elements[i]
                            if element_name and element_name.strip():
                                series_label = SeriesLabel()
                                series_label.v = element_name.strip()
                                series.tx = series_label
                except:
                    pass

            self._apply_safe_mono_series_styling(line_chart, num_elements)

            ws.add_chart(line_chart, line_position)
        
        if chart_config['want_bar']:
            bar_row = chart_positions.get('bar', first_chart_row) + separation_offset
            bar_position = f"{chart_col}{bar_row}"

            num_families = len([f for f in FAMILIES if f in table2.columns]) if not table2.empty else 0
            bar_layout_config = self._calculate_optimal_chart_layout(num_families, "bar")

            bar_chart = BarChart()
            bar_chart.title = "products repartition gaz phase"

            self._apply_ultra_safe_chart_styling(bar_chart, "bar")

            bar_chart.width = bar_layout_config['width']
            bar_chart.height = bar_layout_config['height']

            try:
                from openpyxl.chart.layout import Layout, ManualLayout
                # Ajuster la zone du graphique pour laisser de l'espace aux titres
                bar_chart.layout = Layout(
                    manualLayout=ManualLayout(
                        xMode="edge", yMode="edge",
                        x=0.1,   # Marge gauche pour titre Y (ordonnées)
                        y=0.1,   # Marge haute pour titre principal
                        w=0.75,  # Largeur réduite pour espace légende à droite
                        h=0.7    # Hauteur réduite pour espace titre X (abscisses) en bas
                    )
                )
            except:
                pass

            if num_families <= 1:
                bar_chart.legend = None
            else:
                bar_chart.legend.position = 'r'
                bar_chart.legend.overlay = False

            bar_chart.type = "col"
            bar_chart.grouping = "clustered"

            if not table2.empty:
                filtered_table2 = table2.loc[table2.index.intersection(CARBON_ROWS)]

                if not filtered_table2.empty:
                    family_cols = [f for f in FAMILIES if f in table2.columns]

                    if family_cols:
                        family_col_indices = []
                        for family in family_cols:
                            family_idx = list(table2.columns).index(family)
                            excel_col = table2_col + 1 + family_idx + 1
                            family_col_indices.append(excel_col)

                        if family_col_indices:
                            num_carbon_rows = len(filtered_table2)

                            min_col_families = min(family_col_indices)
                            max_col_families = max(family_col_indices)
                            data_ref = Reference(ws,
                                               min_col=min_col_families,
                                               min_row=table2_row + 1,
                                               max_col=max_col_families,
                                               max_row=table2_row + 1 + num_carbon_rows)
                            bar_chart.add_data(data_ref, titles_from_data=True)

                            carbon_col = table2_col
                            cats = Reference(ws,
                                           min_col=carbon_col,
                                           min_row=table2_row + 2,
                                           max_row=table2_row + 1 + num_carbon_rows)
                            bar_chart.set_categories(cats)

            ws.add_chart(bar_chart, bar_position)
        
        
        return wb


# Test progressif des méthodes
if __name__ == "__main__":
    # ========== CONFIGURATION DU TEST ==========
    # Configurer le nombre d'éléments chimiques à tester
    NOMBRE_ELEMENTS_TEST = 30  # Changer cette valeur pour tester différents nombres d'éléments

    print(f"=== Test progressif ChromeleonOnline - {NOMBRE_ELEMENTS_TEST} éléments ===")

    d = ChromeleonOnline(
        "C:/Users/lucas/Desktop/test")

    # Test 1: _get_data_by_elements
    print("\n1️⃣ Test _get_data_by_elements()")
    try:
        data_by_elements = d._get_data_by_elements()
        print(
            f"✅ Extraction réussie! {len(data_by_elements)} éléments trouvés")
        print(data_by_elements["Methane"].head())
    except Exception as e:
        print(f"❌ Erreur _get_data_by_elements: {e}")
        exit(1)

    # Test 2: get_relative_area_by_injection
    print("\n2️⃣ Test get_relative_area_by_injection()")
    try:
        rel_df = d.get_relative_area_by_injection()
        print("✅ Tableau relatif créé: \n", rel_df)

        # Afficher les premières lignes
        if len(rel_df) > 0:
            print(f"   Première ligne: {rel_df.iloc[0].to_dict()}")
    except Exception as e:
        print(f"❌ Erreur get_relative_area_by_injection: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

    # Test 3: make_summary_tables
    print("\n3️⃣ Test make_summary_tables()")
    try:
        table1, table2 = d.make_summary_tables()
        print(f"✅ Tables de résumé créées:")
        print(f"   Table1 (pics): {table1.shape[0]} lignes")
        print(f"   Table2 (pivot): {table2.shape}")
    except Exception as e:
        print(f"❌ Erreur make_summary_tables: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

    # Test 4: get_graphs_available et génération Excel
    print("\n4️⃣ Test get_graphs_available() et génération Excel")
    try:
        graphs = d.get_graphs_available()
        print(f"✅ Graphiques disponibles: {len(graphs)}")

        # Préparer les métriques pour le fichier Excel
        metrics_wanted = []
        for graph in graphs:
            if graph['available']:
                metric_config = {"name": graph['name']}
                # Pour le graphique temporel, utiliser le nombre d'éléments configuré
                if graph['name'] == "%mass gaz en fonction du temps" and 'chimicalElements' in graph:
                    # Limiter au nombre d'éléments défini dans la configuration de test
                    all_elements = graph['chimicalElements']
                    selected_elements = all_elements[:NOMBRE_ELEMENTS_TEST] if len(all_elements) >= NOMBRE_ELEMENTS_TEST else all_elements

                    metric_config['chimicalElementSelected'] = selected_elements
                    print(
                        f"   - {graph['name']}: ✅ Disponible ({len(selected_elements)} éléments sélectionnés sur {len(all_elements)} disponibles)")
                    print(f"     Éléments testés: {', '.join(selected_elements)}")
                else:
                    print(f"   - {graph['name']}: ✅ Disponible")
                metrics_wanted.append(metric_config)
            else:
                print(f"   - {graph['name']}: ❌ Non disponible")

        # Génération du fichier Excel
        print("\n📊 Génération du fichier Excel...")
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut

        wb = d.generate_workbook_with_charts(
            wb, metrics_wanted, "GC-Online-Test")

        output_file = f"C:/Users/lucas/Desktop/chromeleon_online_test_{NOMBRE_ELEMENTS_TEST}elements.xlsx"
        wb.save(output_file)

        print(f"✅ Fichier Excel créé: {output_file}")
        print(f"📊 Métriques incluses: {[m['name'] for m in metrics_wanted]}")

    except Exception as e:
        print(f"❌ Erreur génération Excel: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

    print("\n🎉 Tous les tests sont passés avec succès!")
    print(f"📁 Fichier Excel généré: {output_file}")
    print("=== Test terminé ===")