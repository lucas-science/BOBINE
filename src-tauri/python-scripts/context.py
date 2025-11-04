import base64
import io
import os
import re
from datetime import datetime
from typing import Union, Optional
from copy import copy  # <<< pour cloner les styles sans DeprecationWarning
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from utils.text_utils import normalize_text, parse_date_value


class ExcelContextData:
    def __init__(self, dir_root: str):
        self.first_file = ""
        if os.path.exists(dir_root):
            files = [
                f for f in os.listdir(dir_root)
                if os.path.isfile(os.path.join(dir_root, f))
                and not f.startswith('.')  
                and not f.startswith('~')
                and not f.startswith('.~lock')
                and f.lower().endswith('.xlsx')
            ]

            if not files:
                raise FileNotFoundError(f"Aucun fichier Excel (.xlsx) valide trouvé dans {dir_root}")

            files.sort()
            self.first_file = os.path.join(dir_root, files[0])
        else:
            raise FileNotFoundError(f"Le répertoire {dir_root} n'existe pas")

        # ouverture du premier fichier Excel
        self.file_path = self.first_file
        self.workbook = load_workbook(self.file_path, data_only=False)
        # première feuille
        self.sheet_name = self.workbook.sheetnames[0]
        self.sheet: Worksheet = self.workbook[self.sheet_name]
    
    def get_masses(self) -> dict[str, Optional[float]]:
        target_labels = {
            "masse recette 1 (kg)": None,
            "masse recette 2 (kg)" : None,
            "masse cendrier (kg)": None,
            "masse injectée (kg)": None,
        }
        data = list(self.sheet.values)
        df = pd.DataFrame(data)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                val = df.iat[i, j]
                if isinstance(val, str) and val.lower() in target_labels.keys():
                    cell_value = df.iat[i, j+1]
                    try:
                        # Convert to float, handling comma as decimal separator
                        if cell_value is not None:
                            target_labels[val.lower()] = float(str(cell_value).replace(',', '.'))
                        else:
                            target_labels[val.lower()] = None
                    except (ValueError, TypeError):
                        # If conversion fails, set to None
                        target_labels[val.lower()] = None 
        
        return target_labels

    def validate(self) -> dict:
        """
        Valide les données du contexte avec des messages d'erreur spécifiques.

        Returns:
            dict: {
                "valid": bool,
                "error_type": str,  # "missing_filename_info" | "missing_experience_data" | "missing_masses" | "invalid_format" | None
                "error_message": str
            }
        """
        try:
            # Test 1: Vérifier les masses
            masses = self.get_masses()
            missing_masses = [k for k, v in masses.items() if v is None]

            if missing_masses:
                return {
                    "valid": False,
                    "error_type": "missing_masses",
                    "error_message": f"Les masses requises sont incomplètes dans le fichier context. Champs manquants: {', '.join(missing_masses)}. Vérifiez que tous les champs de masse sont renseignés."
                }

            # Test 2: Vérifier les informations pour le nom de fichier
            filename_info = self.get_filename_info()
            missing_filename_fields = []

            # Vérifier que les champs critiques ne sont pas aux valeurs par défaut
            if filename_info.get("date") == datetime.now().strftime('%d%m%Y'):
                missing_filename_fields.append("date")
            if filename_info.get("feedstock") == "Unknown":
                missing_filename_fields.append("feedstock")
            if filename_info.get("debit") == "0kgh":
                missing_filename_fields.append("débit plastique")
            if filename_info.get("nb_inducteurs") == "0" or not filename_info.get("temperatures") or filename_info.get("temperatures") == "0":
                missing_filename_fields.append("températures inducteurs")

            if missing_filename_fields:
                return {
                    "valid": False,
                    "error_type": "missing_filename_info",
                    "error_message": f"Les informations pour le nom de fichier sont incomplètes dans le fichier context. Champs manquants ou invalides: {', '.join(missing_filename_fields)}. Vérifiez que les champs date, feedstock, débit plastique et températures des inducteurs sont bien renseignés."
                }

            # Test 3: Vérifier les informations d'expérience
            target_labels = {
                "date": None,
                "heure début": None,
                "heure fin": None
            }

            data = list(self.sheet.values)
            df = pd.DataFrame(data)

            # Parcourir toutes les cellules pour trouver les labels
            for i in range(df.shape[0]):
                for j in range(df.shape[1]):
                    val = df.iat[i, j]
                    if isinstance(val, str):
                        val_clean = val.lower().strip()

                        # Rechercher chaque label cible
                        for key in target_labels.keys():
                            if key in val_clean and target_labels[key] is None:
                                # Récupérer la valeur dans la cellule suivante
                                if j + 1 < df.shape[1]:
                                    next_val = df.iat[i, j + 1]
                                    if next_val is not None and str(next_val).strip():
                                        target_labels[key] = str(next_val).strip()
                                        break

            missing_experience_data = [k for k, v in target_labels.items() if v is None]

            if missing_experience_data:
                return {
                    "valid": False,
                    "error_type": "missing_experience_data",
                    "error_message": f"Les informations d'expérience sont manquantes dans le fichier context. Champs manquants: {', '.join(missing_experience_data)}. Vérifiez que les champs date, heure début et heure fin sont bien renseignés."
                }

            # Si tout est valide
            return {
                "valid": True,
                "error_type": None,
                "error_message": ""
            }

        except Exception as e:
            # Test 4: Format invalide (erreur de lecture Excel, etc.)
            return {
                "valid": False,
                "error_type": "invalid_format",
                "error_message": f"Le format du fichier context n'est pas valide: {str(e)}. Vérifiez qu'il s'agit bien d'un fichier Excel correctement structuré."
            }

    def get_filename_info(self) -> dict:
        """
        Extraction robuste des infos nécessaires pour nommer le fichier :
        Format retourné:
          {
            "date": "DDMMYYYY",
            "feedstock": "LDPE" | "Unknown",
            "debit": "0.73kgh",  # conserve les décimales
            "nb_inducteurs": "3",
            "temperatures": "450450450"
          }
        """
        # Utiliser directement self.sheet qui pointe vers la bonne feuille
        sheet_to_use = self.sheet

        data = list(sheet_to_use.values)
        df = pd.DataFrame(data)
        nrows, ncols = df.shape

        target_info = {
            "date": None,
            "feedstock": None,
            "debit": None,
            "nb_inducteurs": None,
            "temperatures": []
        }

        # Scanner les cellules pour trouver les labels et lire la cellule à droite
        for r in range(nrows):
            for c in range(ncols):
                val = df.iat[r, c]
                if val is None:
                    continue

                norm = normalize_text(val)

                # --- DATE ---
                if target_info["date"] is None and 'date' in norm:
                    if c + 1 < ncols:
                        candidate = df.iat[r, c + 1]
                        parsed = parse_date_value(candidate)
                        if parsed:
                            target_info["date"] = parsed

                # --- FEEDSTOCK ---
                if target_info["feedstock"] is None and ('feedstock' in norm or 'matiere' in norm):
                    if c + 1 < ncols:
                        candidate = df.iat[r, c + 1]
                        if candidate is not None and str(candidate).strip():
                            target_info["feedstock"] = str(candidate).strip().upper()

                # --- DEBIT PLASTIQUE (kg/h) ---
                if target_info["debit"] is None and ('debit' in norm and 'plast' in norm):
                    if c + 1 < ncols:
                        candidate = df.iat[r, c + 1]
                        if candidate is not None:
                            s = str(candidate)
                            m = re.search(r'(\d+[,.]?\d*)', s)
                            if m:
                                num = m.group(1).replace(',', '.')
                                target_info["debit"] = f"{num}kgh"

                # --- NOMBRE D'INDUCTEURS ---
                if target_info["nb_inducteurs"] is None and ('nombre' in norm and 'inducteur' in norm):
                    if c + 1 < ncols:
                        candidate = df.iat[r, c + 1]
                        if candidate is not None:
                            try:
                                num = int(float(str(candidate).replace(',', '.')))
                                target_info["nb_inducteurs"] = str(num)
                            except:
                                pass

        # TEMPERATURES : lire directement B28, C28, D28 (ligne 28, colonnes 2, 3, 4)
        # Ligne 28 = index 27 (0-based), colonnes B=1, C=2, D=3
        if nrows > 27:
            for col_idx in [1, 2, 3]:  # B, C, D
                if col_idx < ncols:
                    cell = df.iat[27, col_idx]
                    if cell is not None:
                        try:
                            val_float = float(str(cell).replace(',', '.'))
                            val_int = int(round(val_float))
                            target_info["temperatures"].append(val_int)
                        except:
                            pass

        # si nb_inducteurs pas trouvé explicitement, on prend la longueur des températures trouvées
        if target_info["nb_inducteurs"] is None:
            target_info["nb_inducteurs"] = str(len(target_info["temperatures"]))

        # si date manquante -> fallback sur date du jour
        if target_info["date"] is None:
            target_info["date"] = datetime.now().strftime('%d%m%Y')

        # fallback feedstock / debit
        if target_info["feedstock"] is None:
            target_info["feedstock"] = "Unknown"

        if target_info["debit"] is None:
            target_info["debit"] = "0kgh"

        temps_concat = "".join(str(int(t)) for t in target_info["temperatures"]) if target_info["temperatures"] else "0"

        result = {
            "date": target_info["date"],
            "feedstock": target_info["feedstock"],
            "debit": target_info["debit"],
            "nb_inducteurs": target_info["nb_inducteurs"],
            "temperatures": temps_concat
        }

        return result

    def get_experience_name(self) -> str:
        """
        Génère le nom de fichier au format V2.
        Format: V2_{date}_{feedstock}_{debit}_{nb_inducteurs}IH_{temperatures}

        Returns:
            str: Nom formaté pour fichier (ex: "V2_11092025_LDPE_0.73kgh_3IH_450450450")
                 ou fallback vers date du jour si informations manquantes
        """
        try:
            info = self.get_filename_info()
            return f"V2_{info['date']}_{info['feedstock']}_{info['debit']}_{info['nb_inducteurs']}IH_{info['temperatures']}"
        except Exception as e:
            # Fallback en cas d'erreur
            today = datetime.now().strftime("%d%m%Y")
            return f"V2_{today}_rapport"

    def get_experience_name_legacy(self) -> str:
        """
        Version legacy de get_experience_name (pour compatibilité si besoin).
        Extrait les informations d'expérience (date, heures) du fichier Excel.

        Returns:
            str: Nom formaté pour fichier (ex: "Rapport_experience_24-juin-25_08h15-16h15")
                 ou fallback vers date du jour si informations manquantes
        """
        target_labels = {
            "date": None,
            "heure début": None,
            "heure fin": None
        }

        data = list(self.sheet.values)
        df = pd.DataFrame(data)
        
        # Parcourir toutes les cellules pour trouver les labels
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                val = df.iat[i, j]
                if isinstance(val, str):
                    val_clean = val.lower().strip()
                    
                    # Rechercher chaque label cible
                    for key in target_labels.keys():
                        if key in val_clean and target_labels[key] is None:
                            # Récupérer la valeur dans la cellule suivante
                            if j + 1 < df.shape[1]:
                                next_val = df.iat[i, j + 1]
                                if next_val is not None and str(next_val).strip():
                                    target_labels[key] = str(next_val).strip()
                                    break
        
        # Construire le nom si toutes les informations sont disponibles
        if all(v is not None for v in target_labels.values()):
            try:
                date_str = str(target_labels["date"]).strip()
                debut_str = str(target_labels["heure début"]).strip()
                fin_str = str(target_labels["heure fin"]).strip()
                
                # Formatter la date proprement
                def format_date(date_str):
                    # Si c'est un timestamp datetime, extraire seulement la partie date
                    if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                        # Format ISO datetime, prendre seulement YYYY-MM-DD
                        date_part = date_str.split()[0]  # Séparer date et heure
                        date_part = date_part.split('T')[0]  # Au cas où format ISO avec T
                        # Convertir en format plus lisible (optionnel)
                        try:
                            from datetime import datetime
                            dt = datetime.strptime(date_part, '%Y-%m-%d')
                            return dt.strftime('%d-%m-%Y')  # Format DD-MM-YYYY
                        except:
                            return date_part.replace('-', '-')
                    else:
                        # Format texte, nettoyer les caractères spéciaux
                        return re.sub(r'[^\w\-]', '-', date_str)
                
                date_clean = format_date(date_str)
                
                # Formatter les heures : extraire HH:MM et convertir en HHhMM
                def format_time(time_str):
                    # Extraire les heures et minutes (format HH:MM)
                    time_match = re.search(r'(\d{1,2}):(\d{2})', time_str)
                    if time_match:
                        hours = time_match.group(1).zfill(2)
                        minutes = time_match.group(2)
                        return f"{hours}h{minutes}"
                    return time_str.replace(":", "h")
                
                debut_clean = format_time(debut_str)
                fin_clean = format_time(fin_str)
                
                return f"Rapport_experience_{date_clean}_{debut_clean}-{fin_clean}"
                
            except Exception:
                pass  # En cas d'erreur, utiliser le fallback
        
        # Fallback : utiliser la date du jour
        today = datetime.now().strftime("%Y-%m-%d")
        return f"Rapport_experience_{today}"
    
    def add_self_sheet_to(self, target_wb: Workbook, new_sheet_name: Optional[str] = None) -> Workbook:
        dst_ws = target_wb.create_sheet(title=self.sheet_name)
        self._copy_sheet(self.sheet, dst_ws)
        return target_wb

    # ---------- export ----------
    def get_as_dataframe(self) -> pd.DataFrame:
        return pd.DataFrame(self.sheet.values)

    def get_as_base64(self) -> str:
        """Encode uniquement la première feuille (copiée dans un mini-workbook) en base64."""
        out_wb = Workbook()
        # nettoyer la sheet par défaut
        if "Sheet" in out_wb.sheetnames and len(out_wb.sheetnames) == 1:
            out_wb.remove(out_wb["Sheet"])
        dst = out_wb.create_sheet(title=self.sheet_name[:31])
        self._copy_sheet(self.sheet, dst)

        bio = io.BytesIO()
        out_wb.save(bio)
        bio.seek(0)
        return base64.b64encode(bio.getvalue()).decode("utf-8")

    # ---------- import / injection ----------
    @staticmethod
    def inject_base64_sheet(
        b64: str,
        target: Union[str, Workbook],
        new_sheet_name: str = "Context",
        save_to: Optional[str] = None,
    ) -> Workbook:
        """
        Injecte la feuille contenue dans `b64` dans un classeur cible.
        - `target`: chemin d'un .xlsx existant OU Workbook openpyxl
        - `new_sheet_name`: nom voulu (sera uniquifié si collision)
        - `save_to`: si `target` est un chemin et que tu veux sauvegarder ailleurs, passe un chemin ici.
        Retourne le Workbook modifié.
        """
        # 1) reconstruire un workbook source depuis le base64
        raw = base64.b64decode(b64)
        src_wb = load_workbook(io.BytesIO(raw), data_only=False)
        src_ws = src_wb.active  # unique feuille dans notre mini-WB

        # 2) ouvrir ou utiliser la cible
        close_after = False
        if isinstance(target, str):
            tgt_wb = load_workbook(target, data_only=False)
            close_after = True
            target_path = target
        else:
            tgt_wb = target
            target_path = None

        # 3) créer une feuille sans collision de nom
        name = new_sheet_name[:31]
        i = 1
        while name in tgt_wb.sheetnames:
            suffix = f"_{i}"
            name = (new_sheet_name[:31 - len(suffix)] + suffix)
            i += 1

        dst_ws = tgt_wb.create_sheet(title=name)
        ExcelContextData._copy_sheet(src_ws, dst_ws)

        # 4) sauvegarde si on a reçu un chemin
        if close_after:
            out_path = save_to or target_path
            tgt_wb.save(out_path)

        return tgt_wb

    # ---------- utilitaire : copie complète d'une feuille ----------
    @staticmethod
    def _copy_sheet(source_ws: Worksheet, target_ws: Worksheet) -> None:
        """
        Copie valeurs, styles, merges, dimensions, freeze panes d'une feuille à l'autre.
        Fix: on utilise enumerate pour obtenir (row_idx, col_idx) et on évite cell.col_idx
        qui n’existe pas sur MergedCell.
        """
        # cellules (valeurs + styles)
        for r_idx, row in enumerate(source_ws.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                t = target_ws.cell(row=r_idx, column=c_idx, value=cell.value)
                # Appliquer styles si présents
                try:
                    # has_style évite d'accéder à des attributs inexistants sur MergedCell
                    if getattr(cell, "has_style", False):
                        if cell.number_format is not None:
                            t.number_format = cell.number_format
                        if cell.font is not None:
                            t.font = copy(cell.font)
                        if cell.fill is not None:
                            t.fill = copy(cell.fill)
                        if cell.alignment is not None:
                            t.alignment = copy(cell.alignment)
                        if cell.border is not None:
                            t.border = copy(cell.border)
                        if cell.protection is not None:
                            t.protection = copy(cell.protection)
                except Exception:
                    # En cas de cellule fusionnée (MergedCell), certains attributs peuvent être absents.
                    # On écrit au minimum la valeur, et on continue.
                    pass

        # merges
        for mrange in getattr(source_ws.merged_cells, "ranges", []):
            target_ws.merge_cells(str(mrange))

        # largeurs colonnes
        for key, dim in source_ws.column_dimensions.items():
            td = target_ws.column_dimensions[key]
            td.width = dim.width
            td.hidden = dim.hidden
            td.outlineLevel = dim.outlineLevel

        # hauteurs lignes
        for idx, dim in source_ws.row_dimensions.items():
            td = target_ws.row_dimensions[idx]
            td.height = dim.height
            td.hidden = dim.hidden
            td.outlineLevel = dim.outlineLevel

        # freeze panes & vue
        target_ws.freeze_panes = source_ws.freeze_panes
        target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
        target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
        target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight

        # largeur minimale si non définie
        if not source_ws.column_dimensions:
            for c in range(1, source_ws.max_column + 1):
                target_ws.column_dimensions[get_column_letter(c)].width = 10


if __name__ == "__main__":

    dir_path = "/home/lucaslhm/Bureau/test"

    try:
        ctx = ExcelContextData(dir_path)
        print(f"✅ Fichier trouvé : {ctx.file_path}")
        print(f"➡️  Feuille active : {ctx.sheet_name}")

        # masses
        masses = ctx.get_masses()
        print("\n📊 Masses extraites :")
        for k, v in masses.items():
            print(f"  - {k} : {v}")

        # validité
        print("\n✔️ Données valides :", ctx.is_valid())

        # nom d'expérience
        print("\n📄 Nom d'expérience généré :", ctx.get_experience_name())

        # dataframe
        print("\n🔎 Aperçu du DataFrame :")
        print(ctx.get_as_dataframe().head())

    except Exception as e:
        print(f"❌ Erreur : {e}")
